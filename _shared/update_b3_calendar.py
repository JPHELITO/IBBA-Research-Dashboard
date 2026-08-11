#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""update_b3_calendar.py — Cronograma de Eventos Corporativos da B3 → Executive Calendar.

Lê a planilha oficial que a B3 publica com as datas PREVISTAS de entrega de ITR e DFP,
cria os earnings da nossa cobertura no calendário da dashboard e publica um arquivo
.ics que o Outlook assina e atualiza sozinho.

Página: https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/acoes/
        consultas/cronograma-de-eventos-corporativos/

O QUE A FONTE É — E O QUE ELA NÃO É
  A B3 informa a data de ENTREGA do ITR/DFP à CVM. É um bom proxy do earnings, mas não
  é a data que o RI anuncia: medido em 2026-08-10, 6 das nossas 7 empresas batiam com o
  cadastro manual e a CSN Mineração divergia (B3 previa 05/08, o cadastro dizia 12/08).
  Por isso a regra número um deste robô é: NUNCA sobrescrever o que foi cadastrado à mão.
  Ele preenche o que está vazio e manda e-mail quando discorda.

  A coluna "Entrega" da planilha é inútil para nós: a B3 a preenche em lote, meses depois
  (em 05/08/2026, 89 empresas já tinham passado da data prevista do 2º tri e NENHUMA
  tinha entrega registrada). Só a coluna "Previsão" é lida.

DUAS ARMADILHAS DA PLANILHA, medidas no arquivo real
  1. as datas vêm em formato MISTO na mesma coluna — 211 textos 'dd/mm/aaaa' e 12 datas
     de verdade do Excel;
  2. tem erro de digitação: '31/09/2026' (31 de setembro não existe) e o ano '226'.
     Um parser ingênuo estoura e derruba o robô inteiro por causa de uma empresa que nem
     é nossa. Aqui a célula ruim é descartada com aviso e a vida segue.

Modos:
  python _shared/update_b3_calendar.py --check      # só diz se saiu arquivo novo
  python _shared/update_b3_calendar.py --dry-run    # mostra tudo o que faria, sem gravar
  python _shared/update_b3_calendar.py --update     # grava no Supabase + publica o .ics
  ... --force     reprocessa mesmo sem mudança      ... --hidden  cria eventos ocultos
  ... --ics-only  só regenera o .ics                ... --peers   inclui peers da B3

Env: SUPABASE_URL, SUPABASE_SERVICE_KEY (obrigatórios p/ gravar), SMTP_USER, SMTP_PASS.
"""
from __future__ import annotations

import argparse
import datetime as dt
import difflib
import hashlib
import io
import json
import os
import re
import sys
import unicodedata
from urllib.parse import urljoin

import requests

HERE = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, HERE)
try:
    import notify
except Exception:                                   # e-mail é acessório, nunca fatal
    notify = None
from ics import Event, build_calendar                # noqa: E402  (depois do sys.path)

PAGE = ("https://www.b3.com.br/pt_br/produtos-e-servicos/negociacao/renda-variavel/"
        "acoes/consultas/cronograma-de-eventos-corporativos/")
UA = {"User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64)",
      "Accept-Language": "pt-BR,pt;q=0.9"}

SOURCE_KEY = "b3_calendar"          # chave em update_log / notify.once
SOURCE_TAG = "b3"                   # valor da coluna `source`
BUCKET = "calendars"
SUPA = os.environ.get("SUPABASE_URL", "").rstrip("/")
SERVICE = os.environ.get("SUPABASE_SERVICE_KEY", "")

# NOME DE PREGÃO (coluna A da planilha) → ticker EXATO como já está no calendário.
# ⚠️ 'VALE' é 'VALE' mesmo, sem sufixo — é o valor usado nos eventos cadastrados à mão
#    (a dashboard usa a ADR de Nova York para a Vale). Não "corrigir" para VALE3.SA.
PREGAO_TO_TICKER = {
    "VALE":          "VALE",
    "CSNMINERACAO":  "CMIN3.SA",
    "USIMINAS":      "USIM5.SA",
    "GERDAU":        "GGBR4.SA",
    "KLABIN S/A":    "KLBN11.SA",
    "SUZANO S.A.":   "SUZB3.SA",
    "IRANI":         "RANI3.SA",
}
# Peers que existem na planilha mas estão fora da cobertura. Só com --peers.
PEERS_TO_TICKER = {"GERDAU MET": "GOAU4.SA", "BRADESPAR": "BRAP3.SA"}

# Nossas empresas que NÃO estão na planilha e nunca estarão — não alarmar por elas.
# CSN é do segmento tradicional (a planilha só traz Novo Mercado / N1 / N2) e a Aura
# é BDR, não entrega ITR na B3. As duas seguem cadastradas à mão.
KNOWN_ABSENT = {"CSN (CSNA3.SA) — segmento tradicional", "Aura (AURA33.SA) — BDR"}

# Índice da coluna de PREVISÃO (0-based) → (tipo, rótulo do trimestre).
# C=DFP(anual) · E=Formulário de Referência (fora: é papelada, não resultado) ·
# G=ITR 1º tri · I=ITR 2º tri · K=ITR 3º tri.
KINDS = {2: ("DFP", 4), 6: ("ITR1", 1), 8: ("ITR2", 2), 10: ("ITR3", 3)}

FIRST_DATA_ROW = 3        # linha 4 da planilha (0-based)
COL_ATUALIZADO = 14       # célula O1 carrega a data de atualização


# ═══════════════════════ 1. buscar a planilha ════════════════════════════════

def fetch_page() -> str | None:
    """HTML da página. Devolve None em falha — nunca levanta."""
    try:
        r = requests.get(PAGE, headers=UA, timeout=60)
        r.raise_for_status()
        return r.text
    except Exception as e:
        print(f"  [B3] falha ao abrir a página ({e}).")
        return None


def find_xlsx_link(html: str) -> str | None:
    """Endereço absoluto da planilha.

    O caminho traz um hash que muda a cada publicação
    (/data/files/06/02/65/3C/A73DF.../Cronograma...xlsx) — não dá para montar a URL,
    tem que achar no HTML. O href é RELATIVO e já vem com %20; urljoin resolve sem
    re-codificar.
    """
    hits = re.findall(r'href="([^"]*data/files/[^"]*\.xlsx?)"', html, re.I)
    hits = [h for h in hits if "cronograma" in h.lower()]
    if len(hits) != 1:
        print(f"  [B3] esperava 1 link de planilha, achei {len(hits)}.")
        _mail("review", f"B3 — layout da página mudou ({len(hits)} links)",
              f"Procurei o link da planilha em {PAGE} e achei {len(hits)} candidatos:\n\n"
              + "\n".join(hits[:10]) +
              "\n\nO robô não mexeu em nada. Provável mudança de layout da B3.",
              period=dt.date.today().strftime("%Y-%m"))
        return None
    return urljoin(PAGE, hits[0])


def download(url: str) -> tuple[bytes | None, str | None]:
    """(conteúdo, cabeçalho Last-Modified)."""
    try:
        r = requests.get(url, headers=UA, timeout=120)
        r.raise_for_status()
        return r.content, r.headers.get("Last-Modified")
    except Exception as e:
        print(f"  [B3] falha ao baixar a planilha ({e}).")
        return None, None


# ═══════════════════════ 2. ler a planilha ═══════════════════════════════════

_BR_DATE = re.compile(r"^\s*(\d{1,2})/(\d{1,2})/(\d{4})\s*$")
_ISO_DATE = re.compile(r"^\s*(\d{4})-(\d{2})-(\d{2})")


def parse_date(v, who: str = "", bad: list | None = None) -> dt.date | None:
    """Uma célula → data, tolerando tudo o que a B3 escreve lá.

    Aceita data de verdade do Excel, texto 'dd/mm/aaaa' (com ou sem apóstrofo à
    frente) e ISO. Devolve None — registrando em `bad` — para vazio, lixo e datas
    impossíveis como o '31/09/2026' que existe no arquivo real.
    """
    if v is None:
        return None
    if isinstance(v, dt.datetime):
        return v.date()
    if isinstance(v, dt.date):
        return v
    s = str(v).strip().strip("'").strip()
    if not s or s in {"-", "--", "N/A", "n/a"}:
        return None
    m = _BR_DATE.match(s)
    if m:
        d, mo, y = int(m.group(1)), int(m.group(2)), int(m.group(3))
        try:
            return dt.date(y, mo, d)
        except ValueError:
            if bad is not None:
                bad.append((who, s))
            return None
    m = _ISO_DATE.match(s)
    if m:
        try:
            return dt.date(*(int(g) for g in m.groups()))
        except ValueError:
            pass
    if bad is not None:
        bad.append((who, s))
    return None


def _norm(s: str) -> str:
    """Normaliza nome de pregão p/ comparar (sem acento, sem espaço duplo, maiúsculo)."""
    s = unicodedata.normalize("NFKD", str(s or ""))
    s = "".join(c for c in s if not unicodedata.combining(c))
    return re.sub(r"\s+", " ", s).strip().upper()


def parse_xlsx(content: bytes) -> tuple[dict, dt.date | None, list]:
    """(dados, data_de_atualizacao, celulas_ruins).

    dados = {NOME DE PREGÃO normalizado: {'ITR3': date, ...}}
    """
    import openpyxl
    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]                     # a planilha tem uma aba só ('ANO')
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return {}, None, []

    bad: list = []
    atualizado = None
    if len(rows[0]) > COL_ATUALIZADO:
        atualizado = parse_date(rows[0][COL_ATUALIZADO], "célula O1")

    out: dict = {}
    for r in rows[FIRST_DATA_ROW:]:
        if not r or not r[0]:
            continue
        nome = _norm(r[0])
        datas = {}
        for idx, (kind, _q) in KINDS.items():
            if idx < len(r):
                d = parse_date(r[idx], f"{r[0]} col{idx}", bad)
                if d:
                    datas[kind] = d
        if datas:
            out[nome] = datas
    return out, atualizado, bad


# ═══════════════════════ 3. Supabase ═════════════════════════════════════════

def _hdr(extra: dict | None = None) -> dict:
    h = {"apikey": SERVICE, "Authorization": f"Bearer {SERVICE}",
         "Content-Type": "application/json"}
    h.update(extra or {})
    return h


def _get(path: str, params: str = "") -> list | None:
    """Lê do PostgREST. Devolve None em FALHA — que é bem diferente de lista vazia.

    ⚠️ Essa distinção é a trava de segurança do robô: se a leitura dos eventos
    existentes falhar e for tratada como "não existe nada", ele passa por cima da
    curadoria manual e duplica tudo. Erro tem que abortar, não seguir calado.
    """
    try:
        r = requests.get(f"{SUPA}/rest/v1/{path}?{params}", headers=_hdr(), timeout=30)
        r.raise_for_status()
        return r.json()
    except Exception as e:
        print(f"  [B3] leitura '{path}' falhou ({e}).")
        return None


def earnings_category_id() -> str | None:
    """Id da categoria de earnings, buscado por `kind` (não fixo no código)."""
    rows = _get("exec_calendar_categories", "select=id,name,kind&kind=eq.earnings")
    if rows:
        return rows[0]["id"]
    for r in (_get("exec_calendar_categories", "select=id,name,kind") or []):
        if re.search(r"earning|conf", r.get("name") or "", re.I):
            return r["id"]
    return None


def last_fingerprint() -> str | None:
    rows = _get("update_log",
                f"select=detail&source=eq.{SOURCE_KEY}&method=eq.fingerprint"
                "&order=id.desc&limit=1")
    return (rows[0].get("detail") if rows else None)


def save_fingerprint(fp: str) -> None:
    try:
        requests.post(f"{SUPA}/rest/v1/update_log",
                      headers=_hdr({"Prefer": "return=minimal"}),
                      data=json.dumps({"source": SOURCE_KEY, "method": "fingerprint",
                                       "detail": fp}), timeout=20)
    except Exception as e:
        print(f"  [B3] não consegui gravar a marca do arquivo ({e}).")


def push_events(rows: list) -> int:
    """Upsert idempotente pela chave natural (source, external_id)."""
    if not rows:
        return 0
    try:
        r = requests.post(
            f"{SUPA}/rest/v1/exec_calendar_events?on_conflict=source,external_id",
            headers=_hdr({"Prefer": "resolution=merge-duplicates,return=minimal,count=exact"}),
            data=json.dumps(rows), timeout=60)
        r.raise_for_status()
        return len(rows)
    except Exception as e:
        print(f"  [B3] gravação falhou ({e}): {getattr(e, 'response', None) and e.response.text}")
        return 0


def hide_event(ev_id: str) -> None:
    """Some do calendário sem apagar — histórico preservado, e o .ics manda CANCELLED."""
    try:
        requests.patch(f"{SUPA}/rest/v1/exec_calendar_events?id=eq.{ev_id}",
                       headers=_hdr({"Prefer": "return=minimal"}),
                       data=json.dumps({"is_visible": False}), timeout=30)
    except Exception as e:
        print(f"  [B3] não consegui ocultar {ev_id} ({e}).")


# ═══════════════════════ 4. montar os eventos ════════════════════════════════

_Q_IN_TITLE = re.compile(r"\|\s*([1-4]Q\d{2})\b", re.I)


def period_label(kind: str, fy: int) -> str:
    """'ITR3' + 2026 → '3Q26'.  'DFP' é o resultado ANUAL do exercício ANTERIOR."""
    q = KINDS_BY_NAME[kind]
    year = fy - 1 if kind == "DFP" else fy
    return f"{q}Q{str(year)[2:]}"


KINDS_BY_NAME = {k: q for (k, q) in KINDS.values()}


def index_existing(rows: list) -> dict:
    """(ticker, '3Q26') → evento. Serve para não pisar no que foi cadastrado à mão."""
    idx = {}
    for e in rows:
        comp = (e.get("company") or "").strip()
        m = _Q_IN_TITLE.search(e.get("title") or "")
        if not comp or not m:
            continue
        idx.setdefault((comp, m.group(1).upper()), []).append(e)
    return idx


def build_events(parsed: dict, fy: int, cat_id: str, existing: list,
                 atualizado: dt.date | None, mapping: dict,
                 hidden: bool = False, today: dt.date | None = None):
    """→ (para_gravar, divergencias, cancelar, ausentes)."""
    today = today or dt.date.today()
    idx = index_existing(existing)
    # Cadeado fechado = você assumiu este evento. Vale tanto para os que você criou
    # do zero quanto para os que vieram da B3 e você corrigiu no admin.
    travados = {e.get("external_id") for e in existing if e.get("locked")}
    upserts, diverge, seen_ext = [], [], set()

    atu = atualizado.strftime("%d/%m/%Y") if atualizado else "?"
    for pregao, ticker in mapping.items():
        datas = parsed.get(_norm(pregao))
        if not datas:
            continue
        for kind, d in sorted(datas.items(), key=lambda kv: kv[1]):
            if d < today:                      # passado não interessa
                continue
            per = period_label(kind, fy)
            ext = f"{SOURCE_TAG}:{ticker}:{fy}:{kind}"
            seen_ext.add(ext)

            # Cadeado fechado: você mandou manter do jeito que está. Nem atualiza,
            # nem reclama — foi uma escolha sua, não uma divergência a resolver.
            if ext in travados:
                continue

            # Já existe algo seu para (empresa, trimestre)? Então o robô não encosta.
            manual = [e for e in idx.get((ticker, per), []) if not e.get("source")]
            if manual:
                m0 = manual[0]
                if m0.get("start_date") != d.isoformat():
                    diverge.append((ticker, per, m0.get("start_date"), d.isoformat()))
                continue

            upserts.append({
                "title":       f"{ticker.split('.')[0]} | {per} Earnings Release (B3)",
                "category_id": cat_id,
                "start_date":  d.isoformat(),
                "all_day":     True,
                "company":     ticker,
                "description": (f"Data prevista de entrega do {('DFP' if kind == 'DFP' else 'ITR')} "
                                f"à CVM, segundo o Cronograma de Eventos Corporativos da B3 "
                                f"(atualizado em {atu}). É uma previsão regulatória — "
                                f"não é a data oficial de divulgação anunciada pelo RI."),
                "links":       [{"label": "Cronograma B3", "url": PAGE}],
                "is_visible":  not hidden,
                "source":      SOURCE_TAG,
                "external_id": ext,
            })

    # Sumiu da planilha? Some do calendário — mas só o que é do robô, só no futuro,
    # e só do exercício que este arquivo cobre. Passado jamais é mexido.
    cancel = [e for e in existing
              if e.get("source") == SOURCE_TAG and e.get("is_visible")
              and not e.get("locked")            # cadeado fechado não some do calendário
              and e.get("external_id") not in seen_ext
              and (e.get("external_id") or "").split(":")[2:3] == [str(fy)]
              and (e.get("start_date") or "") > today.isoformat()]

    ausentes = [p for p in mapping if _norm(p) not in parsed]
    return upserts, diverge, cancel, ausentes


# ═══════════════════════ 5. o arquivo .ics ═══════════════════════════════════

def ics_slug() -> str | None:
    rows = _get("secure_config", "select=value&key=eq.b3_ics_slug")
    return rows[0]["value"] if rows else None


def build_ics(events: list, cancelled: list) -> bytes:
    """Todos os earnings (os seus + os da B3) num calendário só.

    Faz sentido levar os manuais: eles são melhores e cobrem CSN, Aura e as
    estrangeiras, que a B3 não tem. É o calendário que serve de verdade no Outlook.

    `cancelled` são os que sumiram da planilha: continuam no arquivo, marcados como
    cancelados. Se apenas desaparecessem, ficariam fantasmas na agenda de quem assinou.
    """
    gone_ids = {c.get("id") for c in cancelled}
    out = []
    for e in events + cancelled:
        start = dt.date.fromisoformat(e["start_date"])
        end = dt.date.fromisoformat(e["end_date"]) if e.get("end_date") else None
        if not e.get("all_day", True) and e.get("start_time"):
            hh, mm = (e["start_time"].split(":") + ["0"])[:2]
            start = dt.datetime.combine(start, dt.time(int(hh), int(mm)))
            if e.get("end_time"):
                hh2, mm2 = (e["end_time"].split(":") + ["0"])[:2]
                end = dt.datetime.combine(dt.date.fromisoformat(
                    e.get("end_date") or e["start_date"]), dt.time(int(hh2), int(mm2)))
            else:
                end = None
        is_b3 = e.get("source") == SOURCE_TAG
        gone = e.get("id") in gone_ids
        out.append(Event(
            uid=e.get("external_id") or e["id"],
            summary=e.get("title") or "Earnings",
            start=start, end=end,
            sequence=int(e.get("ics_seq") or 0) + (1 if gone else 0),
            # Previsão da B3 entra como "provisório"; o que você curou, como confirmado.
            status="CANCELLED" if gone else ("TENTATIVE" if is_b3 else "CONFIRMED"),
            description=e.get("description") or "",
            categories="Earnings",
        ))
    return build_calendar(out, name="IBBA — Earnings",
                          description="Calendário de resultados da cobertura "
                                      "(curadoria IBBA + previsões do cronograma da B3).")


def publish_ics(data: bytes, slug: str) -> str | None:
    url = f"{SUPA}/storage/v1/object/{BUCKET}/earnings-{slug}.ics"
    try:
        r = requests.post(url, data=data, timeout=60, headers={
            "apikey": SERVICE, "Authorization": f"Bearer {SERVICE}",
            "Content-Type": "text/calendar; charset=utf-8",
            "Cache-Control": "max-age=600", "x-upsert": "true"})
        r.raise_for_status()
        return f"{SUPA}/storage/v1/object/public/{BUCKET}/earnings-{slug}.ics"
    except Exception as e:
        print(f"  [B3] publicação do .ics falhou ({e}).")
        return None


# ═══════════════════════ 6. e-mail e sinais ══════════════════════════════════

def _mail(kind: str, subject: str, body: str, period: str | None = None) -> None:
    if not notify:
        print(f"  [B3] (sem notify) {subject}")
        return
    try:
        if period:
            notify.once(SOURCE_KEY, period, kind, subject, body)
        else:
            notify.send(subject, body)
    except Exception as e:
        print(f"  [B3] e-mail falhou (ignorado): {e}")


def _signal(**kv) -> None:
    """Sinaliza para os passos seguintes do workflow."""
    path = os.environ.get("GITHUB_ENV")
    if not path:
        return
    try:
        with open(path, "a", encoding="utf-8") as f:
            for k, v in kv.items():
                f.write(f"{k}={v}\n")
    except Exception:
        pass


# ═══════════════════════ 7. orquestração ═════════════════════════════════════

def main() -> None:
    ap = argparse.ArgumentParser(description="Cronograma da B3 → Executive Calendar + .ics")
    ap.add_argument("--check", action="store_true", help="só verifica se mudou")
    ap.add_argument("--update", action="store_true", help="grava no Supabase")
    ap.add_argument("--dry-run", action="store_true", help="mostra tudo, não grava nada")
    ap.add_argument("--force", action="store_true", help="reprocessa mesmo sem mudança")
    ap.add_argument("--hidden", action="store_true", help="cria os eventos ocultos")
    ap.add_argument("--ics-only", action="store_true", help="só regenera o .ics")
    ap.add_argument("--peers", action="store_true", help="inclui Gerdau Met e Bradespar")
    a = ap.parse_args()
    write = a.update and not a.dry_run

    if write and (not SUPA or not SERVICE):
        print("  [B3] faltam SUPABASE_URL/SUPABASE_SERVICE_KEY — nada a fazer.")
        _signal(B3_NEW_DATA="false")
        return

    mapping = dict(PREGAO_TO_TICKER)
    if a.peers:
        mapping.update(PEERS_TO_TICKER)

    # ── busca ────────────────────────────────────────────────────────────────
    html = fetch_page()
    if not html:
        _signal(B3_NEW_DATA="false")
        return
    url = find_xlsx_link(html)
    if not url:
        _signal(B3_NEW_DATA="false")
        return
    content, lastmod = download(url)
    if not content:
        _signal(B3_NEW_DATA="false")
        return

    parsed, atualizado, bad = parse_xlsx(content)
    fy = (atualizado or dt.date.today()).year
    sha = hashlib.sha256(content).hexdigest()[:16]
    fp = f"{sha}|{atualizado or '?'}|{lastmod or '?'}"
    print(f"  [B3] planilha de {atualizado or '?'} · {len(parsed)} empresas · marca {fp}")
    if bad:
        print(f"  [B3] {len(bad)} célula(s) de data ilegíveis (descartadas): {bad[:5]}")

    prev = last_fingerprint() if (SUPA and SERVICE) else None
    mudou = a.force or a.dry_run or (prev != fp)
    if not mudou and not a.ics_only:
        print("  [B3] nada mudou desde a última verificação.")
        _signal(B3_NEW_DATA="false")
        return
    if a.check:
        print(f"  [B3] {'MUDOU' if mudou else 'sem mudança'} (anterior: {prev})")
        _signal(B3_NEW_DATA=str(mudou).lower())
        return

    # ── monta ────────────────────────────────────────────────────────────────
    cat_id = earnings_category_id()
    if not cat_id:
        print("  [B3] não achei a categoria de earnings — abortando sem gravar.")
        _mail("review", "B3 — categoria de earnings não encontrada",
              "O robô não achou a categoria com kind='earnings' no calendário. "
              "Nada foi gravado.", period=dt.date.today().strftime("%Y-%m"))
        return

    existing = _get("exec_calendar_events",
                    "select=id,title,company,start_date,end_date,all_day,start_time,"
                    "end_time,description,source,external_id,is_visible,ics_seq,locked"
                    f"&category_id=eq.{cat_id}")
    if existing is None:
        # Sem saber o que já existe, qualquer gravação arrisca duplicar ou passar por
        # cima do que foi cadastrado à mão. Melhor não fazer nada e avisar.
        print("  [B3] não consegui ler os eventos atuais — abortando sem gravar.")
        print("       Se as colunas source/external_id/ics_seq não existem ainda, "
              "rode admin/supabase_b3_calendar.sql no Supabase.")
        _mail("review", "B3 — não consegui ler o calendário atual",
              "O robô não conseguiu listar os eventos já cadastrados e por isso NÃO "
              "gravou nada (gravar às cegas duplicaria eventos ou passaria por cima "
              "da sua curadoria).\n\nCausa mais provável: o arquivo "
              "admin/supabase_b3_calendar.sql ainda não foi rodado no Supabase — sem "
              "ele as colunas source/external_id/ics_seq não existem.",
              period=dt.date.today().strftime("%Y-%m"))
        _signal(B3_NEW_DATA="false")
        return
    upserts, diverge, cancel, ausentes = build_events(
        parsed, fy, cat_id, existing, atualizado, mapping, hidden=a.hidden)

    print(f"\n  == exercício {fy} ==")
    for u in sorted(upserts, key=lambda x: x["start_date"]):
        print(f"    {u['start_date']}  {u['company']:11s} {u['title']}")
    for t, p, meu, b3 in diverge:
        print(f"    ~ {t} {p}: você tem {meu}, a B3 prevê {b3} (não mexi)")
    for c in cancel:
        print(f"    - cancelar: {c.get('title')} ({c.get('start_date')})")
    if not upserts and not diverge and not cancel:
        print("    (nada a fazer)")

    if a.dry_run:
        print("\n  [B3] --dry-run: nada foi gravado.")
        return

    # ── grava ────────────────────────────────────────────────────────────────
    n = 0
    if not a.ics_only:
        n = push_events(upserts)
        for c in cancel:
            hide_event(c["id"])
        save_fingerprint(fp)
        print(f"  [B3] {n} evento(s) gravado(s), {len(cancel)} ocultado(s).")

    # ── .ics ─────────────────────────────────────────────────────────────────
    link = None
    slug = ics_slug()
    if slug:
        fresh = _get("exec_calendar_events",
                     "select=id,title,company,start_date,end_date,all_day,start_time,"
                     "end_time,description,source,external_id,ics_seq,locked"
                     f"&category_id=eq.{cat_id}&is_visible=eq.true"
                     f"&start_date=gte.{(dt.date.today() - dt.timedelta(days=90)).isoformat()}")
        if fresh is None:
            print("  [B3] não consegui reler os eventos — .ics mantido como estava.")
        else:
            link = publish_ics(build_ics(fresh, cancel), slug)
            print(f"  [B3] calendário publicado: {link}")
    else:
        print("  [B3] sem código do .ics em secure_config — rode admin/supabase_b3_calendar.sql.")

    # ── avisos ───────────────────────────────────────────────────────────────
    if ausentes:
        opcoes = {p: difflib.get_close_matches(_norm(p), list(parsed), 3, 0.6) for p in ausentes}
        _mail("review", "B3 — empresa da cobertura sumiu do cronograma",
              "Estes nomes de pregão não apareceram na planilha desta vez:\n\n" +
              "\n".join(f"  {p}   (parecidos no arquivo: {', '.join(v) or '—'})"
                        for p, v in opcoes.items()) +
              "\n\nSe a B3 renomeou o pregão, é só me falar o nome novo que eu ajusto "
              "o robô. Enquanto isso, essa empresa não entra sozinha no calendário.",
              period=str(atualizado or dt.date.today()))

    if diverge:
        _mail("divergence", "B3 — datas diferentes das que você cadastrou",
              "O robô NÃO mexeu em nenhum destes (o seu cadastro manda):\n\n" +
              "\n".join(f"  {t} {p}: você tem {meu}, a B3 prevê {b3}"
                        for t, p, meu, b3 in diverge) +
              "\n\nLembrando: a B3 informa a data de entrega do ITR à CVM, que nem "
              "sempre é a data que o RI anuncia.",
              period=f"{atualizado}|" + ",".join(f"{t}{p}{b3}" for t, p, _m, b3 in diverge))

    if n or cancel:
        corpo = ["Saiu cronograma novo da B3 (atualizado em "
                 f"{atualizado or '?'}).", ""]
        corpo += [f"  + {u['start_date']}  {u['title']}"
                  for u in sorted(upserts, key=lambda x: x["start_date"])]
        corpo += [f"  - cancelado: {c.get('title')} ({c.get('start_date')})" for c in cancel]
        if link:
            corpo += ["", "Calendário para assinar no Outlook "
                          "(Adicionar calendário → Assinar da Web):", f"  {link}"]
        _mail("new_data", f"B3 — cronograma atualizado ({n} evento(s))",
              "\n".join(corpo), period=str(atualizado or dt.date.today()))

    _signal(B3_NEW_DATA="true" if (n or cancel) else "false",
            B3_ATUALIZADO=str(atualizado or ""))


if __name__ == "__main__":
    try:
        main()
    except Exception as e:                       # nunca derruba o workflow
        import traceback
        traceback.print_exc()
        print(f"  [B3] erro inesperado (ignorado): {e}")
    sys.exit(0)
