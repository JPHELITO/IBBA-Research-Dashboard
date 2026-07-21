# -*- coding: utf-8 -*-
"""
montar_pulp.py — monta o pulp_paper.db (Pulp & Paper) AQUI no seu PC.

O QUE FAZ
  1. Baixa o banco ATUAL do GitHub (só leitura, JSON pela API /contents — o
     caminho que funciona na rede do banco; NÃO precisa de token).
  2. Troca SÓ a tabela manual escolhida (IBÁ papel = iba_paper, ou Empapel =
     empapel), lendo a SUA planilha — as tabelas automáticas (SECEX, GACC,
     celulose...) são PRESERVADAS exatamente como estão no ar.
  3. Grava o pulp_paper.db na MESMA pasta deste programa (ex.: G:\\Dashboard).

DEPOIS (publicar — feito por você, pelo navegador, sem script):
  Suba o pulp_paper.db pelo github.com (arrastar e soltar; tem só ~350 KB):
    https://github.com/JPHELITO/IBBA-Research-Dashboard/upload/main/Pulp%20and%20Paper
  → arraste o pulp_paper.db → "Commit changes". A dashboard atualiza em ~1 min.

ONDE DEIXAR A PLANILHA
  Na mesma pasta deste programa (ele procura aqui e nas subpastas) ou em Downloads.

USO
  python montar_pulp.py            (menu)
  python montar_pulp.py iba        (direto: IBÁ papel)
  python montar_pulp.py empapel    (direto: Empapel)
  python montar_pulp.py iba "caminho\\planilha.xlsx"
"""
import base64
import datetime
import json
import os
import sqlite3
import sys
import urllib.request
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

try:
    import openpyxl
except ImportError:
    sys.exit("Falta a biblioteca openpyxl. Instale com:  pip install openpyxl")

VERSION = "2026-07-20"
OWNER, REPO, BRANCH = "JPHELITO", "IBBA-Research-Dashboard", "main"
DB_REL = "Pulp and Paper/pulp_paper.db"           # caminho do banco dentro do repo
OUT_NAME = "pulp_paper.db"                          # nome do arquivo gravado aqui
UA = "ibba-montar-pulp"

SCRIPT_DIR = Path(__file__).resolve().parent
DOWNLOADS = Path.home() / "Downloads"


# ── procurar a planilha (pasta do programa, recursivo, + Downloads) ───────────
def achar_planilha(globs):
    fontes = []
    if os.environ.get("IBBA_BASES_DIR"):
        fontes.append((Path(os.environ["IBBA_BASES_DIR"]), True))
    fontes.append((SCRIPT_DIR, True))
    fontes.append((DOWNLOADS, False))
    cand, vistos = [], set()
    for d, rec in fontes:
        try:
            chave = str(d.resolve()).lower()
        except Exception:
            chave = str(d).lower()
        if chave in vistos or not d.exists():
            continue
        vistos.add(chave)
        for g in globs:
            cand += list(d.rglob(g)) if rec else list(d.glob(g))
    achados = sorted({p for p in cand if not p.name.startswith("~$")},
                     key=lambda p: p.stat().st_mtime, reverse=True)
    if not achados:
        print(f"\n  ✗ Não achei a planilha ({' / '.join(globs)}). Procurei em:")
        for d, _ in fontes:
            if d.exists():
                print(f"      - {d}")
        return None
    if len(achados) == 1:
        return achados[0]
    print("\n  Achei mais de uma planilha — qual é a nova?\n")
    for i, p in enumerate(achados[:6], 1):
        q = datetime.datetime.fromtimestamp(p.stat().st_mtime).strftime("%d/%m/%Y %H:%M")
        print(f"    [{i}] {p.name}   ({q})   [{p.parent}]{'  <- mais recente' if i == 1 else ''}")
    e = input("\n  Número (Enter = a mais recente): ").strip()
    return achados[0] if not e else (achados[int(e) - 1] if e.isdigit() and 1 <= int(e) <= len(achados) else None)


# ── baixar o banco ATUAL pela API /contents (JSON base64; sem token) ──────────
def baixar_db_atual(destino):
    from urllib.parse import quote
    url = f"https://api.github.com/repos/{OWNER}/{REPO}/contents/{quote(DB_REL)}?ref={BRANCH}"
    req = urllib.request.Request(url, headers={
        "User-Agent": UA, "Accept": "application/vnd.github+json",
        "X-GitHub-Api-Version": "2022-11-28"})
    with urllib.request.urlopen(req, timeout=120) as resp:
        d = json.loads(resp.read())
    destino.write_bytes(base64.b64decode(d["content"]))
    return d.get("sha", "")[:7]


# ── helpers de extração (espelham o extractor_pp.py do repo) ──────────────────
def _num(v):
    return float(v) if isinstance(v, (int, float)) else None

def _period(dt):
    return f"{dt.year}-{dt.month:02d}"

def _load(path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)


def extract_iba_paper(path):
    """Aba 'INPUT DATA' — bloco de PAPEL, mensal. Tissue consolidado em 'other'."""
    wb = _load(path); ws = wb["INPUT DATA"]
    rows = [r for r in ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True)]
    wb.close()
    DATE = 8
    P = dict(prod=39, dom=46, exp=53, imp=60, app=67,
             rev=69, rev_latam=70, rev_eu=71, rev_na=72, rev_af=73, rev_as=74, rev_cn=75)
    GRADES = ["total", "packaging", "pw", "newsprint", "tissue", "cardboard", "other"]
    out = []
    for r in rows:
        d = r[DATE]
        if not isinstance(d, datetime.datetime):
            continue
        if _num(r[P["prod"]]) is None and _num(r[P["dom"]]) is None:
            continue
        rec = {"period": _period(d), "year": d.year, "month": d.month}
        for base in ("prod", "dom", "exp", "imp"):
            st = P[base]
            for i, g in enumerate(GRADES):
                rec[f"{base}_{g}"] = _num(r[st + i])
        rec["app_cons"] = _num(r[P["app"]])
        rec["exprev_total"] = _num(r[P["rev"]])
        rec["exprev_latam"] = _num(r[P["rev_latam"]]); rec["exprev_europe"] = _num(r[P["rev_eu"]])
        rec["exprev_namerica"] = _num(r[P["rev_na"]]); rec["exprev_africa"] = _num(r[P["rev_af"]])
        rec["exprev_asia"] = _num(r[P["rev_as"]]);     rec["exprev_china"] = _num(r[P["rev_cn"]])
        out.append(rec)
    for rec in out:                       # tissue + other -> "Others (+Tissue)"
        for m in ("prod", "dom", "exp", "imp"):
            ov, tv = rec.get(f"{m}_other"), rec.get(f"{m}_tissue")
            rec[f"{m}_other"] = None if (ov is None and tv is None) else (ov or 0) + (tv or 0)
            rec.pop(f"{m}_tissue", None)
    return out


def extract_empapel(path):
    """Aba 'CHART DATA' — embarques mensais de papelão ondulado."""
    wb = _load(path); ws = wb["CHART DATA"]
    rows = [r for r in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True)]
    wb.close()
    out = []
    for r in rows:
        d = r[3]
        if not isinstance(d, datetime.datetime):
            continue
        ship = _num(r[6])
        if ship is None:
            continue
        out.append({"period": _period(d), "year": d.year, "month": d.month,
                    "shipments_kton": ship, "working_days": _num(r[7]),
                    "exp_per_day": _num(r[8]), "ltm_shipments": _num(r[13])})
    out.sort(key=lambda x: x["period"])
    return list({x["period"]: x for x in out}.values())   # de-dup por período


# ── bases manuais deste banco ─────────────────────────────────────────────────
_GR = ["total", "packaging", "pw", "newsprint", "cardboard", "other"]
BASES = {
    "iba": {
        "titulo": "IBÁ papel", "tabela": "iba_paper",
        "globs": ["IBÁ*.xlsx", "IBA*.xlsx"],
        "extrair": extract_iba_paper,
        "cols": (["period TEXT", "year INT", "month INT"]
                 + [f"{p}_{g} REAL" for p in ("prod", "dom", "exp", "imp") for g in _GR]
                 + ["app_cons REAL", "exprev_total REAL", "exprev_latam REAL", "exprev_europe REAL",
                    "exprev_namerica REAL", "exprev_africa REAL", "exprev_asia REAL", "exprev_china REAL"]),
    },
    "empapel": {
        "titulo": "Empapel (papelão ondulado)", "tabela": "empapel",
        "globs": ["Empapel*.xlsx"],
        "extrair": extract_empapel,
        "cols": ["period TEXT", "year INT", "month INT", "shipments_kton REAL",
                 "working_days REAL", "exp_per_day REAL", "ltm_shipments REAL"],
    },
}


def resumo(db_path, tabela):
    con = sqlite3.connect(f"file:{db_path}?mode=ro", uri=True)
    r = con.execute(f'SELECT MAX(period), COUNT(*) FROM "{tabela}"').fetchone()
    con.close()
    return r


def trocar_tabela(db_path, tabela, cols, rows):
    """DROP + CREATE + INSERT só da `tabela` — as demais ficam intactas."""
    con = sqlite3.connect(str(db_path)); cur = con.cursor()
    def contar():
        # .fetchall() PRIMEIRO: senão o cur.execute(COUNT) de dentro reseta a iteração
        # da lista de tabelas (mesmo cursor) e o dict sai truncado/errado.
        tabelas = [t for (t,) in cur.execute(
            "SELECT name FROM sqlite_master WHERE type='table'").fetchall()]
        return {t: cur.execute(f'SELECT COUNT(*) FROM "{t}"').fetchone()[0] for t in tabelas}
    antes = contar()
    cur.execute(f'DROP TABLE IF EXISTS "{tabela}"')
    cur.execute(f'CREATE TABLE "{tabela}" ({", ".join(cols)})')
    keys = [c.split()[0] for c in cols]
    cur.executemany(f'INSERT INTO "{tabela}" VALUES ({",".join("?" * len(keys))})',
                    [[r.get(k) for k in keys] for r in rows])
    con.commit()
    depois = contar()
    con.close()
    return antes, depois


def main():
    args = [a for a in sys.argv[1:] if not a.startswith("-")]
    print("=" * 70)
    print(f"  MONTAR pulp_paper.db  (v{VERSION})")
    print("=" * 70)

    chave = args[0].lower() if args else None
    if chave not in BASES:
        print("\n  Qual base manual você quer atualizar?\n")
        for i, (k, b) in enumerate(BASES.items(), 1):
            print(f"    [{i}] {b['titulo']}  (tabela {b['tabela']})")
        e = input("\n  Número: ").strip()
        if not (e.isdigit() and 1 <= int(e) <= len(BASES)):
            sys.exit("\n  Cancelado.")
        chave = list(BASES)[int(e) - 1]
    base = BASES[chave]
    print(f"\n  Base: {base['titulo']}  →  troca a tabela '{base['tabela']}'")

    xlsx = Path(args[1]) if len(args) > 1 else achar_planilha(base["globs"])
    if not xlsx or not xlsx.exists():
        sys.exit("\n  ✗ Planilha não encontrada.")
    print(f"  Planilha: {xlsx.name}  ({xlsx.stat().st_size / 1024:,.0f} KB)")
    print(f"            em {xlsx.parent}")

    # 1) banco atual (preserva as tabelas automáticas)
    print("\n  [1/3] Baixando o pulp_paper.db atual do GitHub (via API, sem token)...")
    out = SCRIPT_DIR / OUT_NAME
    try:
        sha = baixar_db_atual(out)
        print(f"        ok — banco atual (commit {sha})")
    except Exception as e:
        if out.exists():
            print(f"        ⚠ não consegui baixar ({e}). USANDO o pulp_paper.db que já está na pasta.")
        else:
            sys.exit(f"\n  ✗ Não consegui baixar o banco atual e não há um local na pasta.\n     Erro: {e}")

    antes = resumo(out, base["tabela"])
    print(f"        {base['tabela']}: hoje vai até {antes[0]} ({antes[1]:,} linhas)")

    # 2) extrai a planilha e troca a tabela
    print("\n  [2/3] Lendo a planilha e trocando a tabela (as outras ficam intactas)...")
    rows = base["extrair"](str(xlsx))
    if not rows:
        sys.exit("  ✗ Não consegui extrair nenhuma linha da planilha (formato inesperado?).")
    tA, tD = trocar_tabela(out, base["tabela"], base["cols"], rows)
    depois = resumo(out, base["tabela"])

    # 3) resultado
    print("\n  [3/3] Pronto. Resultado:")
    print(f"        {base['tabela']}:  {antes[0]} ({antes[1]:,})  ->  {depois[0]} ({depois[1]:,})")
    mudou = [t for t in tD if t != base["tabela"] and tA.get(t) != tD.get(t)]
    if mudou:
        print(f"        ⚠ ATENÇÃO: outras tabelas mudaram de tamanho ({mudou}) — não era esperado.")
    else:
        print(f"        tabelas automáticas preservadas ({len(tD) - 1}: "
              f"{', '.join(t for t in sorted(tD) if t != base['tabela'])}).")
    if antes == depois:
        print("        ⚠ período e nº de linhas iguais — parece a MESMA planilha já publicada.")
    print(f"\n  Arquivo gravado: {out}  ({out.stat().st_size / 1024:,.0f} KB)")
    print("\n  ── AGORA, PARA PUBLICAR (pelo navegador, sem script) ──────────────────")
    print("  1) Abra:  https://github.com/JPHELITO/IBBA-Research-Dashboard/upload/main/Pulp%20and%20Paper")
    print("  2) Arraste o arquivo pulp_paper.db (o desta pasta) para a página.")
    print('  3) Clique em "Commit changes". A dashboard atualiza em ~1 minuto.')
    print("  " + "-" * 68)


if __name__ == "__main__":
    try:
        main()
    except KeyboardInterrupt:
        print("\n\n  Cancelado.")
