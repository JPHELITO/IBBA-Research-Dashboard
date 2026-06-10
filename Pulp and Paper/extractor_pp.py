#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
extractor_pp.py — Pulp & Paper data extractor for the IBBA Research Dashboard.

Reads the four proprietary source workbooks (default: ~/Downloads) and builds
`pulp_paper.db` (SQLite), read in the browser by pp_dashboard.html via sql.js —
exactly the same architecture as the Steel side (extractor_sm.py / steel_sm.db).

Sources
-------
  IBÁ  — Brazilian Pulp & Paper Association  (production / sales / trade, monthly)
  SECEX — Pulp Foreign Trade by Port         (pulp export volumes/revenue, monthly)
  Empapel — Brazilian Corrugated Paper Assoc.(shipments, domestic demand proxy)
  GACC — China Woodchips Imports             (HW/SW import volumes & prices)

Tables written
--------------
  iba_paper          monthly, by grade (clean, 2008-08 → latest)
  iba_pulp           monthly, by fibre (de-cumulated 2019→; HW/SW/HY pre-2019)
  secex_pulp_port    long format: one row per (period, port)
  calendar           working days per period (Brazilian national holidays)
  company_q          quarterly Klabin / Suzano vs IBÁ (company-level)
  empapel            monthly corrugated-paper shipments
  gacc_woodchips     long format: one row per (period, fibre, country)

Usage
-----
  python extractor_pp.py                 # reads from ~/Downloads, writes ./pulp_paper.db
  python extractor_pp.py --downloads DIR --out FILE
  python extractor_pp.py --iba FILE --secex FILE --empapel FILE --gacc FILE

This is the *bootstrap* path (reads the analyst's Excel). Auto-update from the
primary sources is a later step (mirrors the IABr/INDA roadmap on the Steel side).
"""
import argparse, datetime, os, sqlite3, sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    sys.exit("openpyxl required:  pip install openpyxl")

# ── Default source filenames (in the downloads dir) ──────────────────────────
DEF = {
    "iba":     "IBÁ - Brazilian Pulp & Paper Association.xlsx",
    "secex":   "SECEX - Pulp & Paper Foreign Trade By Port.xlsx",
    "empapel": "Empapel - Brazilian Association of Corrugated Paper.xlsx",
    "gacc":    "GACC - China_Woodchips_Imports_Database.xlsx",
}

# ── helpers ──────────────────────────────────────────────────────────────────
def num(v):
    """Coerce a cell to float, or None for blanks / Excel error strings ('-', #N/A…)."""
    if isinstance(v, (int, float)):
        return float(v)
    return None

def period_of(dt):
    return f"{dt.year}-{dt.month:02d}"

def load(path):
    return openpyxl.load_workbook(path, read_only=True, data_only=True)

# ── Brazilian national holidays → working days per month ─────────────────────
def easter(y):
    a=y%19; b=y//100; c=y%100; d=b//4; e=b%4; f=(b+8)//25; g=(b-f+1)//3
    h=(19*a+b-d-g+15)%30; i=c//4; k=c%4; l=(32+2*e+2*i-h-k)%7
    m=(a+11*h+22*l)//451; mo=(h+l-7*m+114)//31; da=((h+l-7*m+114)%31)+1
    return datetime.date(y, mo, da)

def br_holidays(y):
    e=easter(y)
    fixed=[(1,1),(4,21),(5,1),(9,7),(10,12),(11,2),(11,15),(12,25)]
    days={datetime.date(y,mn,dd) for mn,dd in fixed}
    days.add(e-datetime.timedelta(days=2))    # Good Friday
    days.add(e-datetime.timedelta(days=47))   # Carnival Tuesday
    days.add(e-datetime.timedelta(days=48))   # Carnival Monday
    days.add(e+datetime.timedelta(days=60))   # Corpus Christi
    if y>=2024:
        days.add(datetime.date(y,11,20))      # Black Consciousness (national from 2024)
    return days

_HOL_CACHE={}
def working_days(year, month):
    hol=_HOL_CACHE.setdefault(year, br_holidays(year))
    d=datetime.date(year, month, 1)
    n=0
    while d.month==month:
        if d.weekday()<5 and d not in hol:
            n+=1
        d+=datetime.timedelta(days=1)
    return n

# ── de-cumulation for IBÁ pulp (monthly pre-2019, YTD-cumulative quarters after)─
def decumulate(series):
    """
    series : dict {(year,month): value}  — a single extensive metric.
    Returns dict {(year,month): monthly_value}.

    Per calendar year, auto-detect the reporting mode:
      • MONTHLY    — ≥5 populated months  → use values as-is.
      • CUMULATIVE — ≤4 populated months that strictly increase (quarter-ends hold
                     YTD totals) → take increments between consecutive populated
                     points and spread each evenly over the intervening months.
    """
    out={}
    by_year={}
    for (y,m),v in series.items():
        # 0.0 in the cumulative era means "not reported"; pulp metrics are never a
        # legitimate monthly zero — treat both None and 0 as missing.
        if v:
            by_year.setdefault(y, {})[m]=v
    for y, mv in by_year.items():
        months=sorted(mv)
        vals=[mv[m] for m in months]
        monthly = len(months) >= 5
        increasing = all(vals[i] >= vals[i-1] for i in range(1, len(vals)))
        if monthly or not increasing or len(months) == 1:
            for m in months:
                out[(y,m)] = mv[m]
        else:                                   # cumulative quarters → spread
            prev_m, prev_v = 0, 0.0
            for m in months:
                inc = mv[m] - prev_v
                span = m - prev_m                # months covered by this increment
                if span <= 0: span = 1
                for mm in range(prev_m+1, m+1):
                    out[(y,mm)] = inc/span
                prev_m, prev_v = m, mv[m]
    return out

# ── SECEX port-name normalisation (merge historical URF codes per physical port)─
def norm_port(urf):
    """'0817800 - PORTO DE SANTOS' → 'Santos'. Folds ALF/IRF/PORTO/AEROPORTO variants."""
    s = str(urf or "")
    if " - " in s:
        s = s.split(" - ", 1)[1]
    s = s.upper()
    for pre in ("PORTO DE ", "PORTO DO ", "ALF - ", "ALF-", "IRF - ", "IRF-", "IRF ",
                "AEROPORTO INTERNACIONAL DE ", "AEROPORTO INTERNACIONAL ", "AEROPORTO DE "):
        if s.startswith(pre):
            s = s[len(pre):]
    s = s.strip()
    NAMES = {
        "VITORIA": "Vitória", "SANTOS": "Santos", "RIO GRANDE": "Rio Grande",
        "PARANAGUA": "Paranaguá", "SAO LUIS": "São Luís", "SÃO LUÍS": "São Luís",
        "SAO FRANCISCO DO SUL": "São Francisco do Sul", "ITAJAI": "Itajaí",
        "SALVADOR": "Salvador", "SANTANA": "Santana", "IMBITUBA": "Imbituba",
        "MONTE DOURADO": "Monte Dourado", "ALMEIRIM": "Almeirim",
        "RIO DE JANEIRO": "Rio de Janeiro", "ITAGUAI": "Itaguaí",
        "URUGUAIANA": "Uruguaiana", "FOZ DO IGUACU": "Foz do Iguaçu",
        "FOZ DO IGUAÇU": "Foz do Iguaçu", "CURITIBA": "Curitiba",
        "FORTALEZA": "Fortaleza", "CORUMBA": "Corumbá", "CORUMBÁ": "Corumbá",
        "JAGUARAO": "Jaguarão", "CHUI": "Chuí", "CHUÍ": "Chuí",
        "DIONISIO CERQUEIRA": "Dionísio Cerqueira", "DIONÍSIO": "Dionísio Cerqueira",
        "SAO BORJA": "São Borja", "SÃO BORJA": "São Borja", "PONTA PORA": "Ponta Porã",
    }
    return NAMES.get(s, s.title())

# ═══════════════════════════════════════════════════════════════════════════
#  EXTRACTORS
# ═══════════════════════════════════════════════════════════════════════════
def extract_iba(path):
    """INPUT DATA sheet — clean monthly master for both pulp and paper blocks."""
    wb=load(path); ws=wb["INPUT DATA"]
    rows=[r for r in ws.iter_rows(min_row=12, max_row=ws.max_row, values_only=True)]
    wb.close()

    # column indices (0-based in the values_only tuple) — see header row 11
    DATE=8
    PULP = dict(prod=9, dom=13, exp=17, imp=21, app=25,
                rev=27, rev_latam=28, rev_eu=29, rev_na=30, rev_af=31, rev_as=32, rev_cn=33,
                imprev=34, bal=35)
    PAPER= dict(prod=39, dom=46, exp=53, imp=60, app=67,
                rev=69, rev_latam=70, rev_eu=71, rev_na=72, rev_af=73, rev_as=74, rev_cn=75)

    # ---- PAPER (clean monthly) ----
    paper=[]
    GRADES=["total","packaging","pw","newsprint","tissue","cardboard","other"]
    for r in rows:
        d=r[DATE]
        if not isinstance(d, datetime.datetime): continue
        if num(r[PAPER["prod"]]) is None and num(r[PAPER["dom"]]) is None: continue
        rec={"period":period_of(d),"year":d.year,"month":d.month}
        for base,key in [("prod","prod"),("dom","dom"),("exp","exp"),("imp","imp")]:
            st=PAPER[base]
            for i,g in enumerate(GRADES):
                rec[f"{key}_{g}"]=num(r[st+i])
        rec["app_cons"]=num(r[PAPER["app"]])
        rec["exprev_total"]=num(r[PAPER["rev"]])
        rec["exprev_latam"]=num(r[PAPER["rev_latam"]]); rec["exprev_europe"]=num(r[PAPER["rev_eu"]])
        rec["exprev_namerica"]=num(r[PAPER["rev_na"]]); rec["exprev_africa"]=num(r[PAPER["rev_af"]])
        rec["exprev_asia"]=num(r[PAPER["rev_as"]]);    rec["exprev_china"]=num(r[PAPER["rev_cn"]])
        paper.append(rec)

    # ---- PULP (de-cumulate extensive metrics) ----
    FIB=["total","hw","sw","hy"]
    raw={}   # metric -> {(y,m): value}
    metrics={"prod":PULP["prod"],"dom":PULP["dom"],"exp":PULP["exp"],"imp":PULP["imp"]}
    for key,st in metrics.items():
        for i,f in enumerate(FIB):
            raw[f"{key}_{f}"]={}
    for m in ["app","rev","rev_latam","rev_eu","rev_na","rev_af","rev_as","rev_cn","imprev"]:
        raw[m]={}
    for r in rows:
        d=r[DATE]
        if not isinstance(d, datetime.datetime): continue
        ym=(d.year,d.month)
        for key,st in metrics.items():
            for i,f in enumerate(FIB):
                v=num(r[st+i])
                if v is not None: raw[f"{key}_{f}"][ym]=v
        for mk,ci in [("app",PULP["app"]),("rev",PULP["rev"]),("rev_latam",PULP["rev_latam"]),
                      ("rev_eu",PULP["rev_eu"]),("rev_na",PULP["rev_na"]),("rev_af",PULP["rev_af"]),
                      ("rev_as",PULP["rev_as"]),("rev_cn",PULP["rev_cn"]),("imprev",PULP["imprev"])]:
            v=num(r[ci])
            if v is not None: raw[mk][ym]=v
    dec={k:decumulate(v) for k,v in raw.items()}
    allym=sorted({ym for v in dec.values() for ym in v})
    pulp=[]
    for (y,m) in allym:
        rec={"period":f"{y}-{m:02d}","year":y,"month":m}
        for key in metrics:
            for f in FIB:
                rec[f"{key}_{f}"]=dec[f"{key}_{f}"].get((y,m))
        rec["app_cons"]=dec["app"].get((y,m))
        rec["exprev_total"]=dec["rev"].get((y,m))
        rec["exprev_latam"]=dec["rev_latam"].get((y,m)); rec["exprev_europe"]=dec["rev_eu"].get((y,m))
        rec["exprev_namerica"]=dec["rev_na"].get((y,m)); rec["exprev_africa"]=dec["rev_af"].get((y,m))
        rec["exprev_asia"]=dec["rev_as"].get((y,m));     rec["exprev_china"]=dec["rev_cn"].get((y,m))
        rec["imprev_total"]=dec["imprev"].get((y,m))
        pulp.append(rec)
    return paper, pulp

def extract_company_q(path):
    """Q sheet — quarterly Klabin / Suzano vs IBÁ company-level comparison."""
    wb=load(path); ws=wb["Q"]
    rows=[r for r in ws.iter_rows(min_row=4, max_row=ws.max_row, values_only=True)]
    wb.close()
    out=[]
    for r in rows:
        q=r[32]  # Klabin block 'Q' label (col 33)
        if not isinstance(q,str) or "Q" not in q: continue
        try:
            qq=int(q[0]); yy=2000+int(q[-2:])
        except Exception:
            continue
        rec={"quarter":q,"year":yy,"q":qq,
            # Klabin (cols 34-41 → idx 33-40)
            "klabin_packaging":num(r[33]),"klabin_corr_mi":num(r[34]),"klabin_corr_me":num(r[35]),
            "klabin_sacks_mi":num(r[36]),"klabin_sacks_me":num(r[37]),"klabin_cardboard":num(r[38]),
            "klabin_card_mi":num(r[39]),"klabin_card_me":num(r[40]),
            # Suzano (cols 44-51 → idx 43-50)
            "suzano_total":num(r[43]),"suzano_pw":num(r[44]),"suzano_pw_mi":num(r[45]),
            "suzano_pw_me":num(r[46]),"suzano_cardboard":num(r[47]),"suzano_card_mi":num(r[48]),
            "suzano_card_me":num(r[49]),"suzano_tissue":num(r[50]),
            # IBÁ comparison aggregates (cols 54-58 → idx 53-57)
            "iba_packaging":num(r[53]),"iba_pw":num(r[54]),"iba_cardboard":num(r[55]),
            "iba_tissue":num(r[56])}
        if any(rec[k] is not None for k in rec if k not in ("quarter","year","q")):
            out.append(rec)
    out.sort(key=lambda x:(x["year"],x["q"]))
    return out

def extract_secex(path):
    """DATABASE_PULP — pulp export volumes/revenue by URF (customs port), monthly."""
    wb=load(path); ws=wb["DATABASE_PULP"]
    MO={"01":1,"02":2,"03":3,"04":4,"05":5,"06":6,"07":7,"08":8,"09":9,"10":10,"11":11,"12":12}
    agg={}   # (period, port) -> [vol_kt, rev_mn]
    for r in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        yr,mo,urf,usd,kg=r[0],r[1],r[2],r[3],r[4]
        if not (yr and mo and urf): continue
        try:
            y=int(yr); m=MO.get(str(mo)[:2])
        except Exception:
            continue
        if not m: continue
        vol=num(kg); rev=num(usd)
        if vol is None and rev is None: continue
        port=norm_port(urf)
        per=f"{y}-{m:02d}"
        a=agg.setdefault((per,port),[0.0,0.0,y,m])
        a[0]+=(vol or 0)/1e6     # kg → kton
        a[1]+=(rev or 0)/1e6     # USD → USD mn
    wb.close()
    out=[{"period":p,"year":a[2],"month":a[3],"port":port,
          "volume_ktons":round(a[0],3),"revenue_usd_mn":round(a[1],3)}
         for (p,port),a in agg.items()]
    out.sort(key=lambda x:(x["period"],-x["volume_ktons"]))
    periods=sorted({(x["year"],x["month"]) for x in out})
    cal=[{"period":f"{y}-{m:02d}","year":y,"month":m,"working_days":working_days(y,m)}
         for (y,m) in periods]
    return out, cal

def extract_empapel(path):
    """CHART DATA — monthly corrugated-paper shipments (domestic demand proxy)."""
    wb=load(path); ws=wb["CHART DATA"]
    rows=[r for r in ws.iter_rows(min_row=11, max_row=ws.max_row, values_only=True)]
    wb.close()
    # header row 10 (0-based idx): c4 Date=3, c7 Total Shipments=6, c8 Working Days=7,
    #                c9 Expeditions/WK=8, c14 LTM Shipments=13
    out=[]
    for r in rows:
        d=r[3]
        if not isinstance(d, datetime.datetime): continue
        ship=num(r[6])
        if ship is None: continue
        out.append({"period":period_of(d),"year":d.year,"month":d.month,
                    "shipments_kton":ship,"working_days":num(r[7]),
                    "exp_per_day":num(r[8]),"ltm_shipments":num(r[13])})
    out.sort(key=lambda x:x["period"])
    # de-dup (CHART DATA can repeat the latest month) — keep last seen per period
    seen={x["period"]:x for x in out}
    return list(seen.values())

def extract_gacc(path):
    """WOODCHIP DATA — China woodchip imports, HW/SW by country (long format).

    Layout: two header rows. From col D (idx3) onward, repeating 4-col blocks:
      [USD mn, M tons, USD/ton, <blank>].  Row-1 holds the country label over the
      USD-mn col and the fibre label over the M-tons col.  Volume ('M tons') is in
      millions of bone-dry metric tons (BDMT), matching the OUTPUT sheet.
    """
    wb=load(path); ws=wb["WOODCHIP DATA"]
    rows=[r for r in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True)]
    wb.close()
    if len(rows) < 3: return []
    h1=rows[0]; ncols=max(len(r) for r in rows); data=rows[2:]
    COUNTRY={"Total":"Total","Viet Nam":"Vietnam","Vietnam":"Vietnam","Australia":"Australia",
             "Others":"Others","Other":"Others"}
    out=[]
    for c in range(3, ncols-2, 4):           # need c, c+1, c+2 in range
        country_lbl = h1[c]   if c   < len(h1) else None
        fibre_lbl   = h1[c+1] if c+1 < len(h1) else None
        if not (isinstance(country_lbl,str) and country_lbl.strip()): continue
        country=COUNTRY.get(country_lbl.strip(), country_lbl.strip())
        fib = "HW" if (fibre_lbl and "Hard" in str(fibre_lbl)) else \
              "SW" if (fibre_lbl and "Soft" in str(fibre_lbl)) else str(fibre_lbl or "").strip()
        if fib not in ("HW","SW"): continue
        for r in data:
            dt=r[0]; ym=r[1] if len(r)>1 else None
            if isinstance(dt, datetime.datetime):
                y,m=dt.year,dt.month
            elif isinstance(ym,(int,float)) and ym>190000:
                y,m=int(ym)//100, int(ym)%100
            else:
                continue
            rev=num(r[c]) if c   < len(r) else None
            vol=num(r[c+1]) if c+1 < len(r) else None
            if vol is None and rev is None: continue
            out.append({"period":f"{y}-{m:02d}","year":y,"month":m,"fibre":fib,"country":country,
                        "volume_bdmt":vol,"revenue_usd_mn":rev})
    out.sort(key=lambda x:(x["period"],x["country"],x["fibre"]))
    return out

# ═══════════════════════════════════════════════════════════════════════════
#  DB WRITER
# ═══════════════════════════════════════════════════════════════════════════
def write_db(out_path, paper, pulp, company, secex, cal, empapel, gacc):
    if os.path.exists(out_path):
        os.remove(out_path)
    con=sqlite3.connect(out_path); cur=con.cursor()

    def table(name, rows, cols):
        cur.execute(f"CREATE TABLE {name} ({', '.join(cols)})")
        if rows:
            keys=[c.split()[0] for c in cols]
            ph=",".join("?"*len(keys))
            cur.executemany(f"INSERT INTO {name} VALUES ({ph})",
                            [[r.get(k) for k in keys] for r in rows])

    GR=["total","packaging","pw","newsprint","tissue","cardboard","other"]
    paper_cols=(["period TEXT","year INT","month INT"]
        +[f"{p}_{g} REAL" for p in ("prod","dom","exp","imp") for g in GR]
        +["app_cons REAL","exprev_total REAL","exprev_latam REAL","exprev_europe REAL",
          "exprev_namerica REAL","exprev_africa REAL","exprev_asia REAL","exprev_china REAL"])
    table("iba_paper", paper, paper_cols)

    FB=["total","hw","sw","hy"]
    pulp_cols=(["period TEXT","year INT","month INT"]
        +[f"{p}_{f} REAL" for p in ("prod","dom","exp","imp") for f in FB]
        +["app_cons REAL","exprev_total REAL","exprev_latam REAL","exprev_europe REAL",
          "exprev_namerica REAL","exprev_africa REAL","exprev_asia REAL","exprev_china REAL",
          "imprev_total REAL"])
    table("iba_pulp", pulp, pulp_cols)

    table("company_q", company, ["quarter TEXT","year INT","q INT",
        "klabin_packaging REAL","klabin_corr_mi REAL","klabin_corr_me REAL","klabin_sacks_mi REAL",
        "klabin_sacks_me REAL","klabin_cardboard REAL","klabin_card_mi REAL","klabin_card_me REAL",
        "suzano_total REAL","suzano_pw REAL","suzano_pw_mi REAL","suzano_pw_me REAL",
        "suzano_cardboard REAL","suzano_card_mi REAL","suzano_card_me REAL","suzano_tissue REAL",
        "iba_packaging REAL","iba_pw REAL","iba_cardboard REAL","iba_tissue REAL"])

    table("secex_pulp_port", secex, ["period TEXT","year INT","month INT","port TEXT",
        "volume_ktons REAL","revenue_usd_mn REAL"])
    table("calendar", cal, ["period TEXT","year INT","month INT","working_days INT"])
    table("empapel", empapel, ["period TEXT","year INT","month INT",
        "shipments_kton REAL","working_days REAL","exp_per_day REAL","ltm_shipments REAL"])
    table("gacc_woodchips", gacc, ["period TEXT","year INT","month INT","fibre TEXT","country TEXT",
        "volume_bdmt REAL","revenue_usd_mn REAL"])

    con.commit(); con.close()

# ═══════════════════════════════════════════════════════════════════════════
def main():
    here=Path(__file__).resolve().parent
    ap=argparse.ArgumentParser(description="Build pulp_paper.db from the P&P source workbooks.")
    ap.add_argument("--downloads", default=str(Path.home()/"Downloads"))
    ap.add_argument("--out", default=str(here/"pulp_paper.db"))
    ap.add_argument("--iba"); ap.add_argument("--secex")
    ap.add_argument("--empapel"); ap.add_argument("--gacc")
    a=ap.parse_args()
    dl=Path(a.downloads)
    paths={k:(getattr(a,k) or str(dl/DEF[k])) for k in DEF}

    for k,p in paths.items():
        if not os.path.exists(p):
            sys.exit(f"missing {k} source: {p}")

    print("Reading IBÁ …");      paper, pulp = extract_iba(paths["iba"])
    print("Reading IBÁ (Q) …");  company     = extract_company_q(paths["iba"])
    print("Reading SECEX …");    secex, cal  = extract_secex(paths["secex"])
    print("Reading Empapel …");  empapel     = extract_empapel(paths["empapel"])
    print("Reading GACC …");     gacc        = extract_gacc(paths["gacc"])

    write_db(a.out, paper, pulp, company, secex, cal, empapel, gacc)

    print("\n── pulp_paper.db written ─────────────────────────────")
    def span(rows):
        ps=[r["period"] for r in rows]; return f"{min(ps)} → {max(ps)}" if ps else "—"
    print(f"  iba_paper        {len(paper):>5} rows   {span(paper)}")
    print(f"  iba_pulp         {len(pulp):>5} rows   {span(pulp)}")
    print(f"  company_q        {len(company):>5} rows   "
          f"{company[0]['quarter'] if company else '—'} … {company[-1]['quarter'] if company else '—'}")
    print(f"  secex_pulp_port  {len(secex):>5} rows   {span(secex)}")
    print(f"  empapel          {len(empapel):>5} rows   {span(empapel)}")
    print(f"  gacc_woodchips   {len(gacc):>5} rows   {span(gacc)}")

    # ── self-checks against known values ──
    print("\n── validation ────────────────────────────────────────")
    def find(rows, per): return next((r for r in rows if r["period"]==per), None)
    o=find(paper,"2025-10")
    if o: print(f"  paper 2025-10  prod_total={o['prod_total']} (exp 976) "
                f"app_cons={o['app_cons']} (exp 814)")
    p8=find(pulp,"2008-08")
    if p8: print(f"  pulp  2008-08  prod_total={p8['prod_total']} hw={p8['prod_hw']} "
                 f"sw={p8['prod_sw']} hy={p8['prod_hy']} (exp 1056/885/128/43)")
    may=[r for r in secex if r["period"]=="2026-05"]
    if may:
        tot=sum(r["volume_ktons"] for r in may)
        print(f"  secex 2026-05  total_vol={tot:.1f} kton (exp ~1625.5)  ports={len(may)}")

if __name__=="__main__":
    main()
