#!/usr/bin/env python3
"""
extractor_sm.py — Steel & Mining Dashboard Extractor
=====================================================
Reads source Excel files → populates steel_sm.db (SQLite).

DATA SOURCES
  • IABr  — Brazilian Steel Market Data (.xlsm)
  • SECEX — Steel Foreign Trade Exports (.xlsx)
  • SECEX — Steel Foreign Trade Imports (.xlsx)
  • INDA  — Flat Steel Distributor Data (.xlsm)
  • SECEX — Prediction Analysis / Korea-China CRC-HRC (.xlsx)

USAGE
  pip install pandas openpyxl
  python extractor_sm.py
  python extractor_sm.py --iabr PATH --exp PATH --imp PATH --inda PATH --pred PATH
"""

import sqlite3, sys, argparse
from pathlib import Path
from datetime import datetime

try:
    import pandas as pd
    import openpyxl
except ImportError:
    sys.exit("Missing: pip install pandas openpyxl")

# ── Paths ──────────────────────────────────────────────────────────────────────
HERE      = Path(__file__).parent
DB_PATH   = HERE / "steel_sm.db"
DOWNLOADS = Path.home() / "Downloads"

DEFAULT_IABR = DOWNLOADS / "IABR - Brazilian Steel Market Data.xlsm"
DEFAULT_EXP  = DOWNLOADS / "SECEX - Steel Foreign Trade_Exports.xlsx"
DEFAULT_IMP  = DOWNLOADS / "SECEX - Steel Foreign Trade_Imports.xlsx"
DEFAULT_INDA = DOWNLOADS / "INDA - Flat Steel Distributor Data.xlsm"
DEFAULT_PRED = DOWNLOADS / "SECEX - Prediction Analysis.xlsx"

NOW = datetime.utcnow().isoformat()


# ══════════════════════════════════════════════════════════════════════════════
# HELPERS
# ══════════════════════════════════════════════════════════════════════════════
def _safe_float(v):
    if v is None:
        return None
    try:
        return float(v)
    except (TypeError, ValueError):
        return None

def _period(dt):
    """Convert datetime-like → 'YYYY-MM' string."""
    if isinstance(dt, datetime):
        return dt.strftime("%Y-%m")
    if hasattr(dt, 'strftime'):
        return dt.strftime("%Y-%m")
    s = str(dt)[:7]
    return s if len(s) == 7 else None


# ══════════════════════════════════════════════════════════════════════════════
# DATABASE INIT
# ══════════════════════════════════════════════════════════════════════════════
def init_db(conn):
    conn.executescript("""
    CREATE TABLE IF NOT EXISTS iabr_production (
        period     TEXT PRIMARY KEY,
        year       INTEGER, month INTEGER,
        crude_steel REAL, flat REAL, long_prod REAL,
        semi REAL, slabs REAL, billets REAL, pig_iron REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS iabr_domestic_sales (
        period TEXT PRIMARY KEY,
        year INTEGER, month INTEGER,
        total REAL, flat REAL, long_prod REAL, semi REAL,
        slabs REAL, billets REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS iabr_foreign_market (
        period TEXT PRIMARY KEY,
        year INTEGER, month INTEGER,
        total REAL, flat REAL, long_prod REAL, semi REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS iabr_exports (
        period TEXT PRIMARY KEY,
        year INTEGER, month INTEGER,
        flat_ktons REAL, long_ktons REAL, semi_ktons REAL,
        total_ktons REAL, total_usd_mn REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS iabr_imports (
        period TEXT PRIMARY KEY,
        year INTEGER, month INTEGER,
        flat_ktons REAL, long_ktons REAL, semi_ktons REAL,
        total_ktons REAL, total_usd_mn REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS iabr_consumption (
        period TEXT PRIMARY KEY,
        year INTEGER, month INTEGER,
        total REAL, flat REAL, long_prod REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS secex_exports (
        period   TEXT,
        category TEXT,
        revenue_usd_mn REAL, volume_ktons REAL,
        price_usd_ton  REAL, working_days INTEGER,
        yoy_revenue REAL,   yoy_volume REAL,
        updated_at TEXT,
        PRIMARY KEY (period, category)
    );
    CREATE TABLE IF NOT EXISTS secex_imports (
        period   TEXT,
        category TEXT,
        revenue_usd_mn REAL, volume_ktons REAL,
        price_usd_ton  REAL, working_days INTEGER,
        yoy_revenue REAL,   yoy_volume REAL,
        updated_at TEXT,
        PRIMARY KEY (period, category)
    );
    CREATE TABLE IF NOT EXISTS inda_distribution (
        period TEXT PRIMARY KEY,
        year INTEGER, month INTEGER,
        inventories REAL, inv_months REAL,
        purchases REAL, sales REAL,
        working_days INTEGER, inv_yoy REAL,
        hist_avg REAL, sales_ltm REAL, sales_ma3 REAL,
        updated_at TEXT
    );
    CREATE TABLE IF NOT EXISTS import_prediction (
        period  TEXT,
        product TEXT,
        country TEXT,
        value_usd REAL,
        volume_kg REAL,
        updated_at TEXT,
        PRIMARY KEY (period, product, country)
    );
    CREATE TABLE IF NOT EXISTS pred_exports (
        period     TEXT,
        country    TEXT,
        product    TEXT,
        value_usd  REAL,
        volume_kg  REAL,
        updated_at TEXT,
        PRIMARY KEY (period, country, product)
    );
    """)
    conn.commit()


# ══════════════════════════════════════════════════════════════════════════════
# IABr PARSER (DATABASE sheet — transposed: dates as columns, metrics as rows)
# ══════════════════════════════════════════════════════════════════════════════

# Excel row index (1-based) → field meaning
# Determined from: python -c "... print all row labels"
IABR_ROW_FIELDS = {
    3:  ('prod', 'crude_steel'),
    5:  ('prod', 'flat'),
    6:  ('prod', 'long_prod'),
    7:  ('prod', 'semi'),
    8:  ('prod', 'slabs'),
    9:  ('prod', 'billets'),
    10: ('prod', 'pig_iron'),
    13: ('dom',  'flat'),        # Laminados Planos
    14: ('dom',  'long_prod'),   # Laminados Longos
    15: ('dom',  'semi'),        # Semiacabados subtotal (not stored)
    16: ('dom',  'slabs'),       # Placas → merged into dom flat
    17: ('dom',  'billets'),     # Blocos/Tarugos → merged into dom long
    20: ('for',  'flat'),        # Laminados Planos
    21: ('for',  'long_prod'),   # Laminados Longos
    22: ('for',  'semi'),        # Semiacabados subtotal (not stored)
    23: ('for',  'slabs'),       # Placas → merged into for flat
    24: ('for',  'billets'),     # Blocos/Tarugos → merged into for long
    28: ('exp',  'flat_ktons'),
    29: ('exp',  'long_ktons'),
    30: ('exp',  'semi_ktons'),
    33: ('exp',  'total_ktons'),
    34: ('exp',  'total_usd_mn'),
    37: ('imp',  'flat_ktons'),
    38: ('imp',  'long_ktons'),
    39: ('imp',  'semi_ktons'),
    40: ('imp',  'total_ktons'),
    41: ('imp',  'total_usd_mn'),
    42: ('cons', 'total'),
    43: ('cons', 'flat'),        # Consumo Planos (already includes Slabs per IABr)
    44: ('cons', 'long_prod'),   # Consumo Longos (already includes Billets per IABr)
}

def load_iabr(path, conn):
    print(f"\n[IABr] Loading {path.name} ...")
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    ws = wb['DATABASE']

    all_rows = list(ws.iter_rows(values_only=True))
    if not all_rows:
        print("  [IABr] Empty sheet — skipping.")
        wb.close()
        return

    # Row 0 (Excel row 1) = date headers starting from column B (index 1)
    date_row = all_rows[0]
    dates = []  # list of (col_index, period_str, year, month)
    for ci, cell in enumerate(date_row):
        p = _period(cell)
        if p:
            try:
                dt = datetime.strptime(p, "%Y-%m")
                dates.append((ci, p, dt.year, dt.month))
            except ValueError:
                pass

    print(f"  [IABr] {len(dates)} date columns ({dates[0][1]} to {dates[-1][1]})")

    # Extract metric values per date column
    # data[period] = {field: value}
    data = {p: {'year': y, 'month': m} for (_, p, y, m) in dates}
    col_to_period = {ci: p for (ci, p, _, _) in dates}

    for row_idx_0, row in enumerate(all_rows):
        excel_row = row_idx_0 + 1
        if excel_row not in IABR_ROW_FIELDS:
            continue
        section, field = IABR_ROW_FIELDS[excel_row]
        key = f"{section}_{field}"
        for ci, period in col_to_period.items():
            val = _safe_float(row[ci])
            data[period][key] = val

    wb.close()

    # Insert into tables
    prod_rows, dom_rows, for_rows, exp_rows, imp_rows, cons_rows = [], [], [], [], [], []
    for period, d in sorted(data.items()):
        y, m = d.get('year'), d.get('month')
        # Domestic Sales — guarda as linhas CRUAS do IABr (Planos, Longos, Semiacabados) e
        # também a quebra Placas/Blocos. Quem soma "Flat = Planos + Placas" é a DASHBOARD.
        # ⚠️ Até 2026-08 este extractor já entregava flat JÁ SOMADO com placas e semi=None,
        # enquanto o update_iabr.py (updater ao vivo, que hoje manda na tabela) grava a linha
        # crua — duas definições para a mesma coluna. Alinhado aqui com o updater.
        dom_flat  = d.get('dom_flat')
        dom_long  = d.get('dom_long_prod')
        dom_semi  = d.get('dom_semi')
        dom_slabs = d.get('dom_slabs')
        dom_bill  = d.get('dom_billets')
        dom_total = d.get('dom_total')

        # Foreign Market:
        #   Flat  = Laminados Planos + Placas (Slabs)
        #   Long  = Laminados Longos + Blocos/Tarugos (Ingots, Blooms & Billets)
        _ff = d.get('for_flat');    _fs = d.get('for_slabs')
        _fl = d.get('for_long_prod'); _fb = d.get('for_billets')
        for_flat  = (_ff or 0) + (_fs or 0) if (_ff is not None or _fs is not None) else None
        for_long  = (_fl or 0) + (_fb or 0) if (_fl is not None or _fb is not None) else None
        for_semi  = None  # absorbed into flat and long above
        for_total = (for_flat + for_long) if (for_flat is not None and for_long is not None) else None

        prod_rows.append((
            period, y, m,
            d.get('prod_crude_steel'), d.get('prod_flat'), d.get('prod_long_prod'),
            d.get('prod_semi'), d.get('prod_slabs'), d.get('prod_billets'),
            d.get('prod_pig_iron'), NOW
        ))
        dom_rows.append((period, y, m, dom_total, dom_flat, dom_long, dom_semi,
                         dom_slabs, dom_bill, NOW))
        for_rows.append((period, y, m, for_total, for_flat, for_long, for_semi, NOW))
        exp_rows.append((
            period, y, m,
            d.get('exp_flat_ktons'), d.get('exp_long_ktons'), d.get('exp_semi_ktons'),
            d.get('exp_total_ktons'), d.get('exp_total_usd_mn'), NOW
        ))
        imp_rows.append((
            period, y, m,
            d.get('imp_flat_ktons'), d.get('imp_long_ktons'), d.get('imp_semi_ktons'),
            d.get('imp_total_ktons'), d.get('imp_total_usd_mn'), NOW
        ))
        cons_rows.append((period, y, m, d.get('cons_total'), d.get('cons_flat'), d.get('cons_long_prod'), NOW))

    conn.executemany("INSERT OR REPLACE INTO iabr_production VALUES (?,?,?,?,?,?,?,?,?,?,?)", prod_rows)
    conn.executemany("INSERT OR REPLACE INTO iabr_domestic_sales "
                     "(period,year,month,total,flat,long_prod,semi,slabs,billets,updated_at) "
                     "VALUES (?,?,?,?,?,?,?,?,?,?)", dom_rows)
    conn.executemany("INSERT OR REPLACE INTO iabr_foreign_market VALUES (?,?,?,?,?,?,?,?)", for_rows)
    conn.executemany("INSERT OR REPLACE INTO iabr_exports VALUES (?,?,?,?,?,?,?,?,?)", exp_rows)
    conn.executemany("INSERT OR REPLACE INTO iabr_imports VALUES (?,?,?,?,?,?,?,?,?)", imp_rows)
    conn.executemany("INSERT OR REPLACE INTO iabr_consumption VALUES (?,?,?,?,?,?,?)", cons_rows)
    conn.commit()
    print(f"  [IABr] {len(prod_rows)} months loaded into all IABr tables.")


# ══════════════════════════════════════════════════════════════════════════════
# SECEX EXPORTS / IMPORTS PARSER
# Sheets: SECEX INPUT FLAT / LONGS / SEMI / TOTAL
# Layout: headers at rows 5-6, data from row 7 onwards
# Monthly columns (0-indexed from sheet): 2=Date 3=Rev 4=Vol 5=Price 6=WDays 7=YoYRev 8=YoYVol
# ══════════════════════════════════════════════════════════════════════════════
SECEX_SHEETS = {
    'SECEX INPUT FLAT':  'flat',
    'SECEX INPUT LONGS': 'long',
    'SECEX INPUT SEMI':  'semi',
    'SECEX INPUT TOTAL': 'total',
}

def _parse_secex_sheet(xl, sheet_name, category, table, conn):
    if sheet_name not in xl.sheet_names:
        print(f"    Sheet '{sheet_name}' not found — skipping.")
        return 0

    df = xl.parse(sheet_name, header=None)

    # Find the data header row by looking for a 'Date' cell in col 2
    data_start = None
    for i, row in df.iterrows():
        v = str(row.iloc[2]) if len(row) > 2 else ''
        if 'date' in v.lower() or '1999' in v or '2000' in v:
            # skip — look for actual data rows
            pass
        try:
            if row.iloc[2] is not None and 'date' not in str(row.iloc[2]).lower():
                # Try to parse as date
                dt = pd.to_datetime(row.iloc[2], errors='coerce')
                if dt is not None and not pd.isnull(dt) and dt.year >= 1999:
                    data_start = i
                    break
        except Exception:
            pass

    if data_start is None:
        print(f"    Could not find data in {sheet_name}")
        return 0

    rows_out = []
    for i in range(data_start, len(df)):
        row = df.iloc[i]
        try:
            dt = pd.to_datetime(row.iloc[2], errors='coerce')
            if pd.isnull(dt) or dt.year < 1999 or dt.year > 2030:
                continue
            period = dt.strftime("%Y-%m")
            rev    = _safe_float(row.iloc[3])
            vol    = _safe_float(row.iloc[4])
            price  = _safe_float(row.iloc[5])
            wdays  = _safe_float(row.iloc[6])
            yoy_r  = _safe_float(row.iloc[7])
            yoy_v  = _safe_float(row.iloc[8])
            if rev is None and vol is None:
                continue
            rows_out.append((period, category, rev, vol, price,
                             int(wdays) if wdays is not None else None,
                             yoy_r, yoy_v, NOW))
        except Exception:
            continue

    conn.executemany(
        f"INSERT OR REPLACE INTO {table} "
        "(period,category,revenue_usd_mn,volume_ktons,price_usd_ton,"
        "working_days,yoy_revenue,yoy_volume,updated_at) VALUES (?,?,?,?,?,?,?,?,?)",
        rows_out
    )
    conn.commit()
    return len(rows_out)

def load_secex(path, table, label, conn):
    print(f"\n[SECEX {label}] Loading {path.name} ...")
    xl = pd.ExcelFile(path, engine='openpyxl')
    total = 0
    for sheet_name, category in SECEX_SHEETS.items():
        n = _parse_secex_sheet(xl, sheet_name, category, table, conn)
        print(f"  {sheet_name:25s}: {n} rows ({category})")
        total += n
    print(f"  [SECEX {label}] Total: {total} rows")


# ══════════════════════════════════════════════════════════════════════════════
# INDA PARSER (CHART DATA sheet)
# Columns (0-indexed): 3=Date 6=Inventories 7=InvMonths 8=Purchases 9=Sales
#                      10=InvYoY 11=HistAvg 12=SalesLTM 13=SalesWK 14=SalesMA3
# ══════════════════════════════════════════════════════════════════════════════
def load_inda(path, conn):
    """Read INDA data from INPUT DATA sheet (full history since 2006).
    CHART DATA only has ~73 recent months; INPUT DATA has the complete series.

    INPUT DATA columns (0-indexed):
      8=Date  9=Inventories  10=InvMonths  11=Purchases  12=Sales
      13=WorkingDays  14=InvYoY  15=HistoricalAvg

    sales_ltm and sales_ma3 are not present in INPUT DATA and are computed here.
    """
    print(f"\n[INDA] Loading {path.name} ...")
    wb = openpyxl.load_workbook(path, data_only=True, keep_vba=True)
    ws = wb['INPUT DATA']

    rows = list(ws.iter_rows(values_only=True))

    # Find header row: 'Date' in col I (index 8)
    data_start = None
    for i, row in enumerate(rows):
        if len(row) > 8 and str(row[8]).strip().lower() == 'date':
            data_start = i + 1
            break

    if data_start is None:
        print("  [INDA] Header not found in INPUT DATA.")
        wb.close()
        return

    # Parse raw rows
    raw = []
    for row in rows[data_start:]:
        try:
            dt = row[8]
            if dt is None:
                continue
            if not isinstance(dt, datetime) and not hasattr(dt, 'year'):
                try:
                    dt = datetime.strptime(str(dt)[:10], "%Y-%m-%d")
                except Exception:
                    continue
            p = _period(dt)
            if not p:
                continue
            inv    = _safe_float(row[9])
            inv_mo = _safe_float(row[10])
            purch  = _safe_float(row[11])
            sales  = _safe_float(row[12])
            wdays  = _safe_float(row[13])
            inv_yy = _safe_float(row[14])
            hist   = _safe_float(row[15])
            if inv is None and purch is None and sales is None:
                continue
            dt2 = datetime.strptime(p, "%Y-%m")
            raw.append({
                'period': p, 'year': dt2.year, 'month': dt2.month,
                'inventories': inv, 'inv_months': inv_mo,
                'purchases': purch, 'sales': sales,
                'working_days': int(wdays) if wdays is not None else None,
                'inv_yoy': inv_yy, 'hist_avg': hist,
            })
        except Exception:
            continue

    # Ensure chronological order
    raw.sort(key=lambda r: r['period'])

    # Compute rolling metrics not available in INPUT DATA
    sales_vals = [r['sales'] for r in raw]
    for i, r in enumerate(raw):
        # sales_ltm: 12-month trailing sum (requires 12 observations)
        if i >= 11:
            window = sales_vals[i - 11:i + 1]
            r['sales_ltm'] = sum(v for v in window if v is not None)
        else:
            r['sales_ltm'] = None
        # sales_ma3: 3-month moving average (requires 3 observations)
        if i >= 2:
            window3 = sales_vals[i - 2:i + 1]
            valid3 = [v for v in window3 if v is not None]
            r['sales_ma3'] = round(sum(valid3) / len(valid3), 4) if valid3 else None
        else:
            r['sales_ma3'] = None

    out = [(
        r['period'], r['year'], r['month'],
        r['inventories'], r['inv_months'], r['purchases'], r['sales'],
        r['working_days'], r['inv_yoy'], r['hist_avg'], r['sales_ltm'], r['sales_ma3'],
        NOW
    ) for r in raw]

    conn.executemany(
        "INSERT OR REPLACE INTO inda_distribution "
        "(period,year,month,inventories,inv_months,purchases,sales,"
        "working_days,inv_yoy,hist_avg,sales_ltm,sales_ma3,updated_at) VALUES (?,?,?,?,?,?,?,?,?,?,?,?,?)",
        out
    )
    conn.commit()
    wb.close()
    print(f"  [INDA] {len(out)} months loaded from INPUT DATA ({raw[0]['period']} to {raw[-1]['period']}).")


# ══════════════════════════════════════════════════════════════════════════════
# PREDICTION PARSER (SECEX sheet — CRC/HRC imports from Korea + China)
# Columns: Ano, Mês, NCM, Descrição, SH6, Descrição SH6, País, USD FOB, Kg, DATE, SF6, Value
# We group by DATE + product category (HRC / CRC / Other) + country
# HRC NCMs (SH6 prefix 7208): hot-rolled coil/sheet
# CRC NCMs (SH6 prefix 7209): cold-rolled coil/sheet
# ══════════════════════════════════════════════════════════════════════════════
HRC_SH6 = {
    720890, 720826, 720827, 720837, 720838, 720839, 720853, 720854,
    722540, 722691, 720825, 721190, 722530, 720810, 720836, 721113,
    721114, 721119, 720840,
}
CRC_SH6 = {
    720916, 720917, 720926, 720927, 720915, 720918, 722550, 721129,
    722692, 720990, 721123, 722519, 722619, 720925, 720928,
}

def _classify_sh6(sh6):
    """Exact 6-digit SH6 code match for HRC / CRC classification."""
    try:
        code = int(str(sh6).strip().split('.')[0])
    except (ValueError, TypeError):
        return 'OTHER'
    if code in HRC_SH6:
        return 'HRC'
    if code in CRC_SH6:
        return 'CRC'
    return 'OTHER'

def _classify_country(country):
    c = str(country).lower()
    if 'coreia' in c or 'korea' in c:
        return 'Korea'
    if 'china' in c or 'chin' in c:
        return 'China'
    return 'Other'

def load_korea_exports(xl, conn):
    """KOREA sheet: Korean export data to Brazil.
    Columns (0-indexed): 3=Country(dest), 11=DATE, 12=SF6, 13=Value(USD thousands),
    14=Volume(metric tons → convert ×1000 to kg).
    Filter: Country (iloc[3]) == 'Brazil'.
    """
    if 'KOREA' not in xl.sheet_names:
        print("  [KOREA] KOREA sheet not found — skipping.")
        return
    df = xl.parse('KOREA', header=0)
    agg = {}
    for _, row in df.iterrows():
        try:
            dest = str(row.iloc[3] if not pd.isnull(row.iloc[3]) else '').strip()
            if dest.lower() != 'brazil':
                continue
            dt = pd.to_datetime(row.iloc[11], errors='coerce')
            if pd.isnull(dt):
                continue
            period  = dt.strftime('%Y-%m')
            sf6     = row.iloc[12]
            product = _classify_sh6(sf6)
            if product == 'OTHER':
                continue
            vol_mt  = _safe_float(row.iloc[14]) or 0
            vol_kg  = vol_mt * 1000          # metric tons → kg
            val_usd = (_safe_float(row.iloc[13]) or 0) * 1000  # USD thousands → USD
            key = (period, 'Korea', product)
            if key not in agg:
                agg[key] = [0.0, 0.0]
            agg[key][0] += val_usd
            agg[key][1] += vol_kg
        except Exception:
            continue
    out = [(p, c, pr, v[0], v[1], NOW) for (p, c, pr), v in agg.items()]
    conn.executemany(
        "INSERT OR REPLACE INTO pred_exports "
        "(period,country,product,value_usd,volume_kg,updated_at) VALUES (?,?,?,?,?,?)",
        out)
    conn.commit()
    print(f"  [KOREA] {len(out)} period×product rows (Brazil filter) loaded.")


def load_china_exports(xl, conn):
    """CHINA sheet: Chinese export data to Brazil.
    All rows are already filtered to Brazil (Trading partner = Brazil).
    Columns (0-indexed): 11=DATE, 12=SF6, 13=Value(USD), 14=Volume(kg).
    """
    if 'CHINA' not in xl.sheet_names:
        print("  [CHINA] CHINA sheet not found — skipping.")
        return
    df = xl.parse('CHINA', header=0)
    agg = {}
    for _, row in df.iterrows():
        try:
            dt = pd.to_datetime(row.iloc[11], errors='coerce')
            if pd.isnull(dt):
                continue
            period  = dt.strftime('%Y-%m')
            sf6     = row.iloc[12]
            product = _classify_sh6(sf6)
            if product == 'OTHER':
                continue
            vol_kg  = _safe_float(row.iloc[14]) or 0   # already in kg
            val_usd = _safe_float(row.iloc[13]) or 0
            key = (period, 'China', product)
            if key not in agg:
                agg[key] = [0.0, 0.0]
            agg[key][0] += val_usd
            agg[key][1] += vol_kg
        except Exception:
            continue
    out = [(p, c, pr, v[0], v[1], NOW) for (p, c, pr), v in agg.items()]
    conn.executemany(
        "INSERT OR REPLACE INTO pred_exports "
        "(period,country,product,value_usd,volume_kg,updated_at) VALUES (?,?,?,?,?,?)",
        out)
    conn.commit()
    print(f"  [CHINA] {len(out)} period×product rows loaded.")


def load_prediction(path, conn):
    print(f"\n[PRED] Loading {path.name} ...")
    xl = pd.ExcelFile(path, engine='openpyxl')
    if 'SECEX' not in xl.sheet_names:
        print("  [PRED] SECEX sheet not found.")
        return

    df = xl.parse('SECEX', header=0)
    # Cols: Ano, Mês, Código NCM, Descrição NCM, Código SH6, Descrição SH6,
    #       Países, Valor US$ FOB, Quilograma Líquido, DATE, SF6, Value
    df.columns = [str(c).strip() for c in df.columns]

    # Use the DATE column for period
    date_col = 'DATE'
    usd_col  = 'Valor US$ FOB'
    kg_col   = 'Quilograma Líquido'
    sh6_col  = 'Código SH6'
    country_col = 'Países'

    agg = {}
    for _, row in df.iterrows():
        try:
            dt = pd.to_datetime(row[date_col], errors='coerce')
            if pd.isnull(dt):
                continue
            period  = dt.strftime("%Y-%m")
            product = _classify_sh6(row[sh6_col])
            country = _classify_country(row[country_col])
            usd = _safe_float(row[usd_col]) or 0
            kg  = _safe_float(row[kg_col])  or 0
            key = (period, product, country)
            if key not in agg:
                agg[key] = [0.0, 0.0]
            agg[key][0] += usd
            agg[key][1] += kg
        except Exception:
            continue

    out = [(p, prod, ctry, v[0], v[1], NOW)
           for (p, prod, ctry), v in agg.items()]
    conn.executemany(
        "INSERT OR REPLACE INTO import_prediction "
        "(period,product,country,value_usd,volume_kg,updated_at) VALUES (?,?,?,?,?,?)",
        out
    )
    conn.commit()
    print(f"  [PRED SECEX] {len(out)} period×product×country rows loaded.")

    # Also load Korea and China EXPORT data (leading indicators)
    conn.execute("DELETE FROM pred_exports")
    conn.commit()
    load_korea_exports(xl, conn)
    load_china_exports(xl, conn)


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
def main():
    parser = argparse.ArgumentParser(description="Steel & Mining DB Extractor")
    parser.add_argument('--iabr', default=str(DEFAULT_IABR))
    parser.add_argument('--exp',  default=str(DEFAULT_EXP))
    parser.add_argument('--imp',  default=str(DEFAULT_IMP))
    parser.add_argument('--inda', default=str(DEFAULT_INDA))
    parser.add_argument('--pred', default=str(DEFAULT_PRED))
    args = parser.parse_args()

    conn = sqlite3.connect(DB_PATH)
    init_db(conn)
    print(f"Database: {DB_PATH}")

    iabr = Path(args.iabr)
    exp  = Path(args.exp)
    imp  = Path(args.imp)
    inda = Path(args.inda)
    pred = Path(args.pred)

    if iabr.exists():
        load_iabr(iabr, conn)
    else:
        print(f"\n[IABr] File not found: {iabr}")

    if exp.exists():
        load_secex(exp, 'secex_exports', 'EXP', conn)
    else:
        print(f"\n[SECEX EXP] File not found: {exp}")

    if imp.exists():
        load_secex(imp, 'secex_imports', 'IMP', conn)
    else:
        print(f"\n[SECEX IMP] File not found: {imp}")

    if inda.exists():
        load_inda(inda, conn)
    else:
        print(f"\n[INDA] File not found: {inda}")

    if pred.exists():
        load_prediction(pred, conn)
    else:
        print(f"\n[PRED] File not found: {pred}")

    conn.close()
    print("\nDone.")

if __name__ == '__main__':
    main()
