# -*- coding: utf-8 -*-
"""
atualizar_prediction.py — atualiza o Excel-mestre "SECEX - Prediction Analysis.xlsx"
e publica o resultado na dashboard, em UM comando.

════════════════════════════════════════════════════════════════════════════════
 O QUE ELE FAZ
════════════════════════════════════════════════════════════════════════════════
 Aba SECEX  (importações do Brasil)  → BAIXA SOZINHO do MDIC. Você não faz nada.
 Aba KOREA  (exportações da Coreia)  → lê o arquivo que VOCÊ baixou da alfândega
                                        coreana ("by H.S Code and Country*.xlsx").
 Aba CHINA  (exportações da China)   → lê o CSV que VOCÊ baixou do GACC
                                        ("downloadData*.csv").

 Depois de colar os meses novos, ele:
   • copia as fórmulas auxiliares (colunas L:P) para as linhas novas;
   • manda o Excel recalcular tudo (os gráficos andam sozinhos);
   • recarrega a LINHA PRETA do modelo preditivo (pred_exports) na dashboard;
   • reconstrói o banco web e dá o push  →  a dash atualiza em ~1 min.

 NADA é sobrescrito por engano: mês que já existe na aba é PULADO
 (a não ser que você peça --refazer, para o caso de revisão da fonte).

════════════════════════════════════════════════════════════════════════════════
 COMO USAR (o caminho normal, uma vez por mês)
════════════════════════════════════════════════════════════════════════════════
 1. Baixe a Coreia:  https://tradedata.go.kr  →  "by H.S Code and Country"
    (a busca já salva fica nos 47 códigos certos) → salva em Downloads.
 2. Baixe a China:   https://stats.customs.gov.cn  →  exportar CSV
    (parceiro = Brazil, capítulo 72) → salva "downloadData (N).csv" em Downloads.
 3. Dê dois cliques em  "Atualizar Prediction Analysis.bat".

 É só isso. O SECEX ele busca sozinho.

════════════════════════════════════════════════════════════════════════════════
 MODOS
════════════════════════════════════════════════════════════════════════════════
   python atualizar_prediction.py --conferir
       Só mostra o que cada aba tem hoje e o que está faltando. NÃO escreve nada.

   python atualizar_prediction.py
       Faz tudo: preenche o Excel + publica na dashboard.

   python atualizar_prediction.py --so-excel      só preenche o Excel
   python atualizar_prediction.py --so-dash       só recarrega a dash (Excel já pronto)
   python atualizar_prediction.py --sem-push      faz tudo, menos o git push

   --excel  CAMINHO     usa outro Excel-mestre (padrão: o mais novo do Downloads)
   --coreia CAMINHO     usa outro arquivo da Coreia
   --china  CAMINHO     usa outro CSV da China   (pode repetir: --china a.csv --china b.csv)
   --refazer 2026-07    apaga esse mês das abas e regrava (use quando a fonte revisar)

════════════════════════════════════════════════════════════════════════════════
 SEGURANÇA
════════════════════════════════════════════════════════════════════════════════
 • Antes de mexer, salva uma cópia em Downloads\_backup_prediction\.
 • Escreve pelo PRÓPRIO Excel (COM), então gráficos, formatação e vínculos
   externos da planilha ficam intactos — nada de biblioteca reescrevendo o arquivo.
 • Se o Excel-mestre estiver aberto, ele avisa e para (não corrompe).
"""
import argparse
import csv
import glob
import io
import os
import re
import shutil
import subprocess
import sys
import tempfile
from collections import defaultdict
from datetime import datetime
from pathlib import Path

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

HERE = Path(__file__).resolve().parent
DOWNLOADS = Path.home() / "Downloads"
BACKUP_DIR = DOWNLOADS / "_backup_prediction"
SM_DIR = HERE / "Steel and Mining"

MDIC_NCM = "https://balanca.economia.gov.br/balanca/bd/comexstat-bd/ncm"
MDIC_TAB = "https://balanca.economia.gov.br/balanca/bd/tabelas"
UA = {"User-Agent": "Mozilla/5.0"}

MESES_PT = ["", "Janeiro", "Fevereiro", "Março", "Abril", "Maio", "Junho",
            "Julho", "Agosto", "Setembro", "Outubro", "Novembro", "Dezembro"]

XL_UP = -4162          # constante xlUp do Excel
ABAS = ("SECEX", "KOREA", "CHINA")


# ═══════════════════════════════════════════════════════════════════════════════
# utilidades pequenas
# ═══════════════════════════════════════════════════════════════════════════════
def titulo(t):
    print("\n" + "─" * 78 + f"\n{t}\n" + "─" * 78)


def num(x, padrao=0.0):
    """Converte texto do tipo '   1,234.5 ' em número. Devolve `padrao` se não der."""
    if x is None:
        return padrao
    if isinstance(x, (int, float)):
        return float(x)
    s = str(x).replace(",", "").replace('"', "").strip()
    if not s or s in ("-", "?"):
        return padrao
    try:
        return float(s)
    except ValueError:
        return padrao


def todos_que_casam(padrao_glob, pasta=None):
    """Arquivos que casam com o padrão, do mais novo p/ o mais velho (sem os '~$')."""
    pasta = pasta or DOWNLOADS
    achados = [Path(p) for p in glob.glob(str(pasta / padrao_glob))
               if not Path(p).name.startswith("~$")]
    return sorted(achados, key=lambda p: p.stat().st_mtime, reverse=True)


def mais_novo(padrao_glob, pasta=None):
    """Arquivo mais recente que casa com o padrão (ignora os temporários '~$')."""
    achados = todos_que_casam(padrao_glob, pasta)
    return achados[0] if achados else None


def _relata_ja_tinha(vistos, ja_na_aba, refazer=None):
    """Uma linha curta dizendo quantos meses do arquivo a aba já tinha (sem listar 67)."""
    rep = sorted(m for m in vistos if m in ja_na_aba and m != refazer)
    if not rep:
        return
    if len(rep) <= 3:
        print(f"    já estavam na aba (pulei): {', '.join(rep)}")
    else:
        print(f"    já estavam na aba (pulei): {len(rep)} meses, {rep[0]} → {rep[-1]}")


def melhor_por_mes(por_arquivo):
    """
    Recebe {arquivo: {mes: [linhas]}} e devolve ({mes: [linhas]}, [avisos]).

    REGRA: para cada mês, vale UM arquivo só — o que trouxe MAIS linhas (empate:
    o mais recente). Nunca soma dois arquivos no mesmo mês.
    Por quê: é comum sobrar na pasta um download parcial do mesmo mês (uma consulta
    que você fez com filtro mais estreito). Somar contaria duas vezes; escolher o
    mais completo acerta. Os descartados aparecem como aviso, nunca em silêncio.
    """
    escolhido, avisos = {}, []
    meses = {m for d in por_arquivo.values() for m in d}
    for mes in sorted(meses):
        cands = [(len(d[mes]), f) for f, d in por_arquivo.items() if mes in d]
        cands.sort(key=lambda t: (t[0], t[1].stat().st_mtime), reverse=True)
        n, ganhador = cands[0]
        escolhido[mes] = por_arquivo[ganhador][mes]
        if len(cands) > 1:
            outros = ", ".join(f"{f.name} ({k} linhas)" for k, f in cands[1:])
            avisos.append(f"{mes}: usei {ganhador.name} ({n} linhas); "
                          f"ignorei os parciais → {outros}")
    return escolhido, avisos


def esta_aberto(xlsx: Path) -> bool:
    """Excel cria um '~$arquivo.xlsx' enquanto o arquivo está aberto."""
    return (xlsx.parent / ("~$" + xlsx.name)).exists()


def ym(ano, mes):
    return f"{int(ano):04d}-{int(mes):02d}"


# ═══════════════════════════════════════════════════════════════════════════════
# 1. LER O ESTADO ATUAL DO EXCEL  (o que cada aba já tem)
# ═══════════════════════════════════════════════════════════════════════════════
def ler_estado(xlsx: Path) -> dict:
    """
    Devolve, por aba: meses presentes, última linha com dado, e o universo de
    códigos que a aba usa (o escopo é DITADO PELA PLANILHA — nunca por mim).
    """
    import openpyxl
    wb = openpyxl.load_workbook(xlsx, data_only=True, read_only=True)
    est = {}

    # ── SECEX: A=Ano, B='07. Julho', C=NCM, E=SH6, F=desc SH6, D=desc NCM
    ws = wb["SECEX"]
    meses, sh6, desc_ncm, desc_sh6, ultima = set(), set(), {}, {}, 1
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r[0] is None:
            continue
        ultima = i
        meses.add(ym(r[0], str(r[1])[:2]))
        s = str(r[4]).zfill(6)
        sh6.add(s)
        desc_ncm[str(r[2]).zfill(8)] = r[3]
        desc_sh6[s] = r[5]
    est["SECEX"] = dict(meses=meses, ultima=ultima, sh6=sh6,
                        desc_ncm=desc_ncm, desc_sh6=desc_sh6)

    # ── KOREA: A='2026.07', B=HS6
    ws = wb["KOREA"]
    meses, hs, ultima = set(), set(), 1
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r[0] is None:
            continue
        ultima = i
        p = str(r[0]).strip()
        if "." in p:
            meses.add(p.replace(".", "-"))
        hs.add(str(r[1]).strip().zfill(6))
    est["KOREA"] = dict(meses=meses, ultima=ultima, hs=hs)

    # ── CHINA: A=202607, B=código 8 dígitos
    ws = wb["CHINA"]
    meses, ultima = set(), 1
    for i, r in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if r[0] is None:
            continue
        ultima = i
        p = str(r[0]).strip()
        if len(p) == 6 and p.isdigit():
            meses.add(f"{p[:4]}-{p[4:]}")
    est["CHINA"] = dict(meses=meses, ultima=ultima)

    wb.close()
    return est


# ═══════════════════════════════════════════════════════════════════════════════
# 2. SECEX — busca automática no MDIC
# ═══════════════════════════════════════════════════════════════════════════════
def _baixa_csv_mdic(url, cache_nome=None):
    """Baixa um CSV do MDIC (latin-1, ';'). Guarda em cache do dia p/ não repetir."""
    import requests
    import urllib3
    urllib3.disable_warnings()
    cache = None
    if cache_nome:
        cdir = Path(tempfile.gettempdir()) / "ibba_mdic_cache"
        cdir.mkdir(exist_ok=True)
        cache = cdir / f"{datetime.now():%Y%m%d}_{cache_nome}"
        if cache.exists() and cache.stat().st_size > 1000:
            print(f"    (cache de hoje: {cache.name})")
            return cache.read_bytes()
    r = requests.get(url, verify=False, timeout=600, headers=UA)
    r.raise_for_status()
    if cache:
        cache.write_bytes(r.content)
    return r.content


def secex_meses_disponiveis(ano):
    """Meses já publicados pelo MDIC no ano (lê só 2 colunas do arquivo do ano)."""
    import pandas as pd
    b = _baixa_csv_mdic(f"{MDIC_NCM}/IMP_{ano}.csv", f"IMP_{ano}.csv")
    df = pd.read_csv(io.StringIO(b.decode("latin-1", "replace")), sep=";", dtype=str,
                     usecols=["CO_ANO", "CO_MES"], low_memory=False)
    return sorted({ym(a, m) for a, m in zip(df["CO_ANO"].str.strip(), df["CO_MES"].str.strip())})


def secex_linhas(ano, meses_alvo, sh6_set, desc_ncm, desc_sh6):
    """
    Monta as linhas da aba SECEX (A..I) para os meses pedidos, direto do MDIC.
    Reproduz EXATAMENTE o formato da aba: importações, todos os países, agregado
    por NCM × país, só os SH6 que a planilha já usa.
    """
    import pandas as pd
    b = _baixa_csv_mdic(f"{MDIC_NCM}/IMP_{ano}.csv", f"IMP_{ano}.csv")
    df = pd.read_csv(io.StringIO(b.decode("latin-1", "replace")), sep=";", dtype=str,
                     usecols=["CO_ANO", "CO_MES", "CO_NCM", "CO_PAIS", "KG_LIQUIDO", "VL_FOB"],
                     low_memory=False)
    df["CO_NCM"] = df["CO_NCM"].str.strip().str.zfill(8)
    df["CO_MES"] = df["CO_MES"].str.strip().str.zfill(2)
    df["per"] = df["CO_ANO"].str.strip() + "-" + df["CO_MES"]
    df = df[df["per"].isin(meses_alvo) & df["CO_NCM"].str[:6].isin(sh6_set)]
    if df.empty:
        return []

    # nomes de país / descrições — MDIC é a fonte; a planilha é o fallback
    pais = pd.read_csv(io.StringIO(_baixa_csv_mdic(f"{MDIC_TAB}/PAIS.csv", "PAIS.csv")
                                   .decode("latin-1", "replace")), sep=";", dtype=str)
    pais.columns = [c.strip().strip('"') for c in pais.columns]
    pmap = {str(a).strip().zfill(3): str(b_).strip()
            for a, b_ in zip(pais["CO_PAIS"], pais["NO_PAIS"])}

    tn = pd.read_csv(io.StringIO(_baixa_csv_mdic(f"{MDIC_TAB}/NCM.csv", "NCM.csv")
                                 .decode("latin-1", "replace")), sep=";", dtype=str)
    tn.columns = [c.strip().strip('"') for c in tn.columns]
    nmap = {str(a).strip().zfill(8): str(b_).strip()
            for a, b_ in zip(tn["CO_NCM"], tn["NO_NCM_POR"])}

    ts = pd.read_csv(io.StringIO(_baixa_csv_mdic(f"{MDIC_TAB}/NCM_SH.csv", "NCM_SH.csv")
                                 .decode("latin-1", "replace")), sep=";", dtype=str)
    ts.columns = [c.strip().strip('"') for c in ts.columns]
    smap = {str(a).strip().zfill(6): str(b_).strip()
            for a, b_ in zip(ts["CO_SH6"], ts["NO_SH6_POR"])}

    df["pais"] = df["CO_PAIS"].str.strip().str.zfill(3).map(pmap).fillna("Outros")
    df["kg"] = pd.to_numeric(df["KG_LIQUIDO"], errors="coerce").fillna(0)
    df["usd"] = pd.to_numeric(df["VL_FOB"], errors="coerce").fillna(0)
    g = (df.groupby(["CO_ANO", "CO_MES", "CO_NCM", "pais"], as_index=False)[["usd", "kg"]]
           .sum().sort_values(["CO_ANO", "CO_MES", "CO_NCM", "pais"]))

    linhas = []
    for _, r in g.iterrows():
        ncm = r["CO_NCM"]
        s6 = ncm[:6]
        mes_i = int(r["CO_MES"])
        linhas.append([
            int(r["CO_ANO"]),
            f"{mes_i:02d}. {MESES_PT[mes_i]}",
            int(ncm),
            nmap.get(ncm) or desc_ncm.get(ncm) or "",
            int(s6),
            smap.get(s6) or desc_sh6.get(s6) or "",
            r["pais"],
            int(round(r["usd"])),
            int(round(r["kg"])),
        ])
    return linhas


# ═══════════════════════════════════════════════════════════════════════════════
# 3. KOREA — lê o arquivo baixado da alfândega coreana
# ═══════════════════════════════════════════════════════════════════════════════
def _korea_um_arquivo(caminho: Path, hs_conhecidos: set):
    """Um 'by H.S Code and Country*.xlsx' → ({mes: [linhas A..I]}, HS ignorados)."""
    import openpyxl
    wb = openpyxl.load_workbook(caminho, data_only=True, read_only=True)
    ws = wb[wb.sheetnames[0]]
    linhas_brutas = list(ws.iter_rows(values_only=True))
    wb.close()
    # o cabeçalho fica na linha 5 (acima vêm título, filtros e a unidade) e a
    # linha logo abaixo é o 'TOTAL' — as duas coisas têm de ser puladas
    cab = next((i for i, r in enumerate(linhas_brutas)
                if r and str(r[0]).strip() == "Period"), None)
    if cab is None:
        return {}, set()

    por_mes, fora = defaultdict(list), set()
    for r in linhas_brutas[cab + 1:]:
        if not r or r[0] is None:
            continue
        per = str(r[0]).strip()
        if per in ("", "TOTAL") or "." not in per:
            continue
        hs = str(r[1]).strip().zfill(6)
        if hs not in hs_conhecidos:
            fora.add(hs)
            continue
        por_mes[per.replace(".", "-")].append([
            per, hs, r[2], r[3],
            num(r[4]), num(r[5]), num(r[6]), num(r[7]), num(r[8]),
        ])
    return dict(por_mes), fora


def korea_linhas(caminhos, hs_conhecidos: set, ja_na_aba=frozenset(), refazer=None):
    """
    Lê os arquivos da alfândega coreana e devolve {mes: [linhas A..I]}.
    Só os HS que a aba KOREA já usa — o escopo é o da planilha, não o meu.
    Mês que a aba já tem é descartado aqui (a não ser o --refazer); mês repetido
    em mais de um arquivo: vale o mais completo.
    """
    por_arquivo, fora_total, descartados, vistos = {}, set(), 0, set()
    for p in caminhos:
        try:
            d, fora = _korea_um_arquivo(Path(p), hs_conhecidos)
        except Exception:
            d, fora = {}, set()
        fora_total |= fora
        if not d:
            descartados += 1
            continue
        vistos |= set(d)
        d = {m: v for m, v in d.items() if m not in ja_na_aba or m == refazer}
        if d:
            por_arquivo[Path(p)] = d
    if descartados:
        print(f"    ({descartados} arquivo(s) sem o cabeçalho 'Period' — ignorados)")
    if fora_total:
        print(f"    (ignorei {len(fora_total)} códigos que a aba KOREA não usa: "
              f"{', '.join(sorted(fora_total)[:6])}{'…' if len(fora_total) > 6 else ''})")
    _relata_ja_tinha(vistos, ja_na_aba, refazer)
    if not por_arquivo:
        return {}
    escolhido, avisos = melhor_por_mes(por_arquivo)
    for av in avisos:
        print(f"    ⚠ {av}")
    return escolhido


# ═══════════════════════════════════════════════════════════════════════════════
# 4. CHINA — lê o(s) CSV do GACC
# ═══════════════════════════════════════════════════════════════════════════════
def _china_um_arquivo(p: Path, sh6_set):
    """
    Um CSV do GACC → {mes: [linhas A..J]}, ou {} se o arquivo não serve.

    Duas escolhas que parecem detalhe mas não são:

    1) NÃO agrega. A aba CHINA guarda as linhas BRUTAS do download — o mesmo código
       aparece várias vezes (uma por regime aduaneiro / província), só que essas duas
       colunas não são coladas. Somar daria o total certo, mas mudaria a cara da aba;
       o modelo usa SUMIFS, então tanto faz para a conta e é melhor ficar fiel.

    2) Lê em latin-1, não em UTF-8. O arquivo do GACC vem em GBK; o Excel decodifica
       como latin-1 na hora de colar, e é por isso que as 3.988 linhas que já estão lá
       trazem "w¡Ý600mm" em vez de "w≥600mm". Lendo em latin-1 as linhas novas saem
       iguaizinhas às antigas. (A coluna de descrição não entra em nenhuma fórmula.)
    """
    por_mes = defaultdict(list)
    with open(p, encoding="latin-1", newline="") as f:
        rd = csv.reader(f)
        cab = next(rd, None)
        if not cab:
            return {}
        H = [str(c).strip().lower() for c in cab]

        def idx(nome):
            return H.index(nome) if nome in H else None
        iD, iC, iN = idx("date of data"), idx("commodity code"), idx("commodity")
        iPC, iP = idx("trading partner code"), idx("trading partner")
        iQ, iU = idx("quantity"), idx("unit")
        iSQ, iSU, iV = idx("supplementary quantity"), idx("supplementary unit"), idx("us dollar")
        if None in (iD, iC, iP, iQ, iV):
            return {}
        for r in rd:
            if len(r) <= iV:
                continue
            per = str(r[iD]).strip()
            cod = str(r[iC]).strip()
            if not (len(per) == 6 and per.isdigit()) or not cod.isdigit():
                continue
            if str(r[iP]).strip() != "Brazil":
                continue
            if cod[:6] not in sh6_set:
                continue
            por_mes[f"{per[:4]}-{per[4:]}"].append([
                int(per), int(cod),
                r[iN] if iN is not None else "",
                int(num(r[iPC], 410)) if iPC is not None else 410,
                "Brazil",
                int(round(num(r[iQ]))),
                r[iU] if iU is not None else "Kilogram",
                int(round(num(r[iSQ]))) if iSQ is not None else 0,
                r[iSU] if iSU is not None else "?",
                int(round(num(r[iV]))),
            ])
    return dict(por_mes)


def china_linhas(caminhos, sh6_set, ja_na_aba=frozenset(), refazer=None):
    """
    Lê os downloadData*.csv do GACC (só parceiro 'Brazil', só os SH6 da planilha)
    e agrega o detalhe por regime aduaneiro / província, que a aba CHINA não tem.
    Mês que a aba já tem é descartado aqui (a não ser o --refazer); quando o mesmo
    mês aparece em mais de um arquivo, vale o mais completo.
    """
    por_arquivo, descartados, vistos = {}, 0, set()
    for p in caminhos:
        try:
            d = _china_um_arquivo(Path(p), sh6_set)
        except Exception:
            d = {}
        if not d:
            descartados += 1
            continue
        vistos |= set(d)
        d = {m: v for m, v in d.items() if m not in ja_na_aba or m == refazer}
        if d:
            por_arquivo[Path(p)] = d
    if descartados:
        print(f"    ({descartados} arquivo(s) não eram CSV de aço do GACC — ignorados)")
    _relata_ja_tinha(vistos, ja_na_aba, refazer)
    if not por_arquivo:
        return {}
    escolhido, avisos = melhor_por_mes(por_arquivo)
    for av in avisos:
        print(f"    ⚠ {av}")
    return escolhido


# ═══════════════════════════════════════════════════════════════════════════════
# 5. ESCREVER NO EXCEL (pelo próprio Excel, via COM)
# ═══════════════════════════════════════════════════════════════════════════════
class Planilha:
    """Abre o Excel-mestre, escreve blocos no fim das abas e recalcula."""

    def __init__(self, caminho: Path):
        import time
        import win32com.client as win32
        import pythoncom
        self.caminho = caminho
        # Dispatch (late binding) e não EnsureDispatch: o EnsureDispatch precisa gerar
        # os wrappers da type library e falha ("can not automate the makepy process")
        # quando o Excel ainda está subindo. E o Excel recusa chamadas enquanto está
        # ocupado ("Call was rejected by callee") — daí as tentativas com pausa.
        ultimo = None
        for tentativa in range(6):
            try:
                pythoncom.CoInitialize()
                # DispatchEx e não Dispatch: o Dispatch se PENDURA num Excel que já esteja
                # rodando — inclusive num processo órfão travado (acontece direto no Windows,
                # sobra um EXCEL.EXE sem janela) e aí tudo falha. O DispatchEx sobe um
                # processo só nosso, e no fim a gente fecha só ele. O seu Excel fica em paz.
                self.xl = win32.DispatchEx("Excel.Application")
                for prop, val in (("Visible", False), ("DisplayAlerts", False),
                                  ("AskToUpdateLinks", False), ("EnableEvents", False)):
                    try:
                        setattr(self.xl, prop, val)
                    except Exception:
                        pass          # conveniências: se o Excel recusar, seguimos assim mesmo
                self._separador_americano()
                # UpdateLinks=0: não tenta resolver os vínculos externos mortos da planilha
                self.wb = self.xl.Workbooks.Open(str(caminho), UpdateLinks=0, ReadOnly=False)
                return
            except Exception as e:            # pywintypes.com_error e afins
                ultimo = e
                try:
                    self.xl.Quit()
                except Exception:
                    pass
                if tentativa:
                    print(f"    (Excel ocupado, tentativa {tentativa + 1}/6…)")
                time.sleep(1.5 * (tentativa + 1))
        raise SystemExit(f"Não consegui falar com o Excel: {ultimo}\n"
                         "Feche as janelas do Excel abertas e rode de novo.")

    def _separador_americano(self):
        """
        ⚠ NÃO TIRE ISTO. É o passo que impede a planilha de se estragar sozinha.

        As colunas auxiliares da aba KOREA usam =NUMBERVALUE(E) e a maior parte das
        47 mil linhas de E está gravada como TEXTO ("   37.6"). O NUMBERVALUE sem
        argumentos usa o separador decimal do WINDOWS — que aqui é a VÍRGULA. Nesse
        caso ele lê o ponto como separador de MILHAR:

            NUMBERVALUE("37.6")   ->  376      (10× a mais)
            NUMBERVALUE("48.0")   ->  480      (10× a mais)
            NUMBERVALUE("1,234.5")->  #VALOR!
            NUMBERVALUE(37.6)     ->  37,6     (certo — quando é NÚMERO, não texto)

        A planilha foi montada com separador americano, e por isso os valores
        guardados nela estão certos. Mas basta um recálculo geral com a configuração
        brasileira para todas essas linhas virarem 10× — foi o que aconteceu no teste.
        Aqui a gente força ponto/vírgula NA NOSSA instância do Excel (que é só nossa e
        é fechada no fim), então o recálculo reproduz os mesmos números de sempre.
        A configuração do SEU Excel não é tocada.
        """
        try:
            self.xl.UseSystemSeparators = False
            self.xl.DecimalSeparator = "."
            self.xl.ThousandsSeparator = ","
        except Exception as e:
            print(f"    ⚠ não consegui fixar o separador decimal ({e}). "
                  "Vou seguir SEM recálculo geral, para não estragar as linhas antigas.")
            self._sem_rebuild = True

    def ultima_linha(self, aba):
        ws = self.wb.Worksheets(aba)
        return ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row

    def apagar_mes(self, aba, linhas_idx):
        """
        Apaga linhas inteiras. Junta os índices em BLOCOS contíguos e apaga bloco a
        bloco, de baixo p/ cima (se fosse uma a uma, cada exclusão moveria as de
        baixo — e 700 chamadas ao Excel demoram muito).
        """
        ws = self.wb.Worksheets(aba)
        idx = sorted(set(linhas_idx))
        blocos, ini, ant = [], idx[0], idx[0]
        for i in idx[1:]:
            if i == ant + 1:
                ant = i
                continue
            blocos.append((ini, ant))
            ini = ant = i
        blocos.append((ini, ant))
        for a, b in reversed(blocos):
            ws.Range(ws.Rows(a), ws.Rows(b)).Delete()
        return len(blocos)

    # Colunas que precisam entrar como TEXTO, não como número.
    # Só a KOREA: "2026.07" e "721030" são RÓTULOS — a coluna auxiliar os lê com
    # LEFT()/RIGHT(). Se o Excel os converter em número, "2026.07" vira 2026,07 e o
    # rótulo perde o sentido. (Nas abas SECEX e CHINA o próprio arquivo já guarda
    # esses campos como número — conferido linha a linha contra o original.)
    COLUNAS_TEXTO = {"KOREA": (1, 2)}

    def anexar(self, aba, linhas, col_ini=1):
        """Cola o bloco no fim e ESTENDE as fórmulas auxiliares (L:P) para as linhas novas."""
        if not linhas:
            return 0
        ws = self.wb.Worksheets(aba)
        ult = ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row
        n, ncols = len(linhas), len(linhas[0])
        for c in self.COLUNAS_TEXTO.get(aba, ()):
            ws.Range(ws.Cells(ult + 1, c), ws.Cells(ult + n, c)).NumberFormat = "@"
        alvo = ws.Range(ws.Cells(ult + 1, col_ini), ws.Cells(ult + n, col_ini + ncols - 1))
        alvo.Value = tuple(tuple(x for x in lin) for lin in linhas)
        # fórmulas auxiliares: copia a última linha boa (colunas L..P = 12..16)
        origem = ws.Range(ws.Cells(ult, 12), ws.Cells(ult, 16))
        destino = ws.Range(ws.Cells(ult + 1, 12), ws.Cells(ult + n, 16))
        origem.Copy(destino)
        self.xl.CutCopyMode = False
        return n

    def linhas_do_mes(self, aba, mes):
        """Índices das linhas de um mês (p/ o --refazer)."""
        ws = self.wb.Worksheets(aba)
        ult = ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row
        if ult < 2:
            return []
        col_a = ws.Range(ws.Cells(2, 1), ws.Cells(ult, 1)).Value
        col_b = ws.Range(ws.Cells(2, 2), ws.Cells(ult, 2)).Value
        achados = []
        for i, (a, b) in enumerate(zip(col_a, col_b), start=2):
            va = a[0] if isinstance(a, tuple) else a
            vb = b[0] if isinstance(b, tuple) else b
            if va is None:
                continue
            if aba == "SECEX":
                chave = ym(va, str(vb)[:2]) if vb else None
            elif aba == "KOREA":
                chave = str(va).strip().replace(".", "-")
            else:
                s = str(int(va)) if isinstance(va, float) else str(va).strip()
                chave = f"{s[:4]}-{s[4:6]}" if len(s) == 6 else None
            if chave == mes:
                achados.append(i)
        return achados

    def amostra_koreia(self, quantos=400):
        """
        Lê uma amostra de (Volume digitado, Volume calculado) da aba KOREA.
        Serve para conferir, ANTES de salvar, que o recálculo não mexeu em linha
        antiga: a coluna O tem de continuar igual à coluna E, linha por linha.
        """
        ws = self.wb.Worksheets("KOREA")
        ult = ws.Cells(ws.Rows.Count, 1).End(XL_UP).Row
        if ult < 2:
            return {}
        passo = max(1, (ult - 1) // quantos)
        linhas = list(range(2, ult + 1, passo))
        amostra = {}
        for i in linhas:
            amostra[i] = (ws.Cells(i, 5).Value, ws.Cells(i, 15).Value)
        return amostra

    @staticmethod
    def _confere_amostra(amostra):
        """Devolve a lista de linhas em que o Volume calculado não bate com o digitado."""
        ruins = []
        for i, (e, o) in amostra.items():
            if e is None:
                continue
            try:
                ev = float(str(e).replace(",", "").strip())
            except (TypeError, ValueError):
                continue
            if isinstance(o, str) or o is None:          # #VALOR! e afins
                ruins.append((i, e, o))
            elif abs(ev - float(o)) > max(1e-6, abs(ev) * 1e-9):
                ruins.append((i, e, o))
        return ruins

    def fechar(self, salvar=True):
        try:
            if salvar:
                if not getattr(self, "_sem_rebuild", False):
                    self.xl.CalculateFullRebuild()
                else:
                    self.xl.Calculate()
                ruins = self._confere_amostra(self.amostra_koreia())
                if ruins:
                    salvar = False
                    print(f"\n⚠ ABORTEI SEM SALVAR: {len(ruins)} linhas da aba KOREA saíram do "
                          "recálculo com o volume errado (problema do separador decimal).")
                    for i, e, o in ruins[:5]:
                        print(f"    linha {i}: digitado {e!r} → calculado {o!r}")
                    print("  Seu arquivo NÃO foi alterado.")
                else:
                    self.wb.Save()
            return salvar
        finally:
            for f in (lambda: self.wb.Close(SaveChanges=False), self.xl.Quit):
                try:
                    f()
                except Exception:
                    pass


# ═══════════════════════════════════════════════════════════════════════════════
# 6. PUBLICAR NA DASHBOARD
# ═══════════════════════════════════════════════════════════════════════════════
def rodar(cmd, cwd=None):
    print("   $ " + " ".join(str(c) for c in cmd[:3]) + (" …" if len(cmd) > 3 else ""))
    p = subprocess.run(cmd, cwd=str(cwd or HERE), capture_output=True, text=True,
                       encoding="utf-8", errors="replace")
    saida = (p.stdout or "") + (p.stderr or "")
    for lin in saida.strip().splitlines()[-14:]:
        print("     " + lin)
    if p.returncode != 0:
        raise SystemExit(f"Falhou (código {p.returncode}). Veja o log acima.")
    return saida


def publicar(excel: Path, sem_push=False):
    titulo("PUBLICANDO NA DASHBOARD")
    py = sys.executable
    print("  1/4  linha preta (pred_exports) a partir do Excel")
    rodar([py, str(SM_DIR / "reload_pred_exports.py"), "--pred", str(excel)])
    print("  2/4  portão de tamanho do banco")
    rodar([py, str(HERE / "_shared" / "check_db_size.py"), str(SM_DIR / "steel_sm.db")])
    print("  3/4  banco web (o que o cliente baixa)")
    rodar([py, str(SM_DIR / "build_web_db.py")])
    if sem_push:
        print("  4/4  --sem-push: parei antes do git. Publique quando quiser.")
        return
    print("  4/4  git commit + push")
    arqs = ["Steel and Mining/steel_sm.db", "Steel and Mining/steel_sm_web.db.gz"]
    rodar(["git", "add"] + arqs)
    st = subprocess.run(["git", "diff", "--staged", "--quiet"], cwd=str(HERE))
    if st.returncode == 0:
        print("     nada mudou no banco — nada a publicar.")
        return
    import sqlite3
    ult = sqlite3.connect(SM_DIR / "steel_sm.db").execute(
        "SELECT MAX(period) FROM pred_exports").fetchone()[0]
    rodar(["git", "commit", "-m", f"data: modelo preditivo ate {ult} (pred_exports)"])
    rodar(["git", "push"])
    print("\n  ✅ publicado. A dashboard atualiza em ~1 minuto.")


# ═══════════════════════════════════════════════════════════════════════════════
# PRINCIPAL
# ═══════════════════════════════════════════════════════════════════════════════
def main():
    ap = argparse.ArgumentParser(add_help=True, description=__doc__,
                                 formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument("--conferir", action="store_true", help="só mostra o que falta; não escreve")
    ap.add_argument("--so-excel", action="store_true", help="só preenche o Excel")
    ap.add_argument("--so-dash", action="store_true", help="só publica na dashboard")
    ap.add_argument("--sem-push", action="store_true", help="faz tudo menos o git push")
    ap.add_argument("--excel", help="caminho do Excel-mestre")
    ap.add_argument("--coreia", help="caminho do arquivo da Coreia")
    ap.add_argument("--china", action="append", help="caminho de CSV do GACC (pode repetir)")
    ap.add_argument("--refazer", help="apaga e regrava esse mês (ex.: 2026-07)")
    a = ap.parse_args()

    # ── achar o Excel-mestre ──────────────────────────────────────────────────
    excel = Path(a.excel) if a.excel else mais_novo("SECEX - Prediction Analysis*.xlsx")
    if not excel or not excel.exists():
        raise SystemExit("Não achei o 'SECEX - Prediction Analysis*.xlsx' no Downloads.\n"
                         "Baixe o arquivo ou passe --excel CAMINHO.")
    titulo("ATUALIZAR PREDICTION ANALYSIS")
    print(f"Excel-mestre : {excel.name}")
    print(f"              ({datetime.fromtimestamp(excel.stat().st_mtime):%d/%m/%Y %H:%M}, "
          f"{excel.stat().st_size/1e6:.1f} MB)")

    if a.so_dash:
        publicar(excel, a.sem_push)
        return

    est = ler_estado(excel)
    print("\nO que cada aba tem hoje:")
    for aba in ABAS:
        m = sorted(est[aba]["meses"])
        print(f"  {aba:6s} {len(m):3d} meses   {m[0]} → {m[-1]}   (última linha {est[aba]['ultima']})")

    # ── juntar o que há de novo em cada fonte ────────────────────────────────
    titulo("PROCURANDO MESES NOVOS")
    novos = {}

    # SECEX (automático)
    print("SECEX  — buscando no MDIC (balanca.economia.gov.br)…")
    ja = est["SECEX"]["meses"]
    ano_atual = max(int(m[:4]) for m in ja)
    disp = []
    for ano in range(ano_atual, datetime.now().year + 1):
        try:
            disp += secex_meses_disponiveis(ano)
        except Exception as e:
            print(f"    (ano {ano}: {type(e).__name__} — {str(e)[:80]})")
    alvo = sorted(m for m in disp if m not in ja or m == a.refazer)
    if alvo:
        print(f"    meses a inserir: {', '.join(alvo)}")
        linhas = []
        for ano in sorted({m[:4] for m in alvo}):
            linhas += secex_linhas(int(ano), [m for m in alvo if m[:4] == ano],
                                   est["SECEX"]["sh6"], est["SECEX"]["desc_ncm"],
                                   est["SECEX"]["desc_sh6"])
        novos["SECEX"] = linhas
        print(f"    {len(linhas)} linhas prontas.")
    else:
        print(f"    nada novo (MDIC também está em {max(disp) if disp else '?'}).")

    # KOREA (arquivo baixado)
    print("KOREA  — lendo o(s) arquivo(s) da alfândega coreana…")
    kfs = ([Path(a.coreia)] if a.coreia
           else todos_que_casam("by H.S Code and Country*.xlsx"))
    if not kfs:
        print("    nenhum 'by H.S Code and Country*.xlsx' no Downloads — pulando a Coreia.")
    else:
        print(f"    {len(kfs)} arquivo(s), mais novo: {kfs[0].name}")
        pm = korea_linhas(kfs, est["KOREA"]["hs"], est["KOREA"]["meses"], a.refazer)
        if pm:
            alvo = sorted(pm)
            novos["KOREA"] = [l for m in alvo for l in pm[m]]
            print(f"    meses a inserir: {', '.join(alvo)} → {len(novos['KOREA'])} linhas.")
        else:
            print("    nada novo nesses arquivos.")

    # CHINA (CSV baixado)
    print("CHINA  — lendo o(s) CSV do GACC…")
    cfs = ([Path(c) for c in a.china] if a.china
           else todos_que_casam("downloadData*.csv"))
    if not cfs:
        print("    nenhum 'downloadData*.csv' no Downloads — pulando a China.")
    else:
        print(f"    {len(cfs)} arquivo(s), mais novo: {cfs[0].name}")
        pm = china_linhas(cfs, est["SECEX"]["sh6"], est["CHINA"]["meses"], a.refazer)
        if pm:
            alvo = sorted(pm)
            novos["CHINA"] = [l for m in alvo for l in pm[m]]
            print(f"    meses a inserir: {', '.join(alvo)} → {len(novos['CHINA'])} linhas.")
        else:
            print("    nada novo nesses arquivos.")

    # ── resumo / saída antecipada ────────────────────────────────────────────
    titulo("RESUMO")
    if not novos:
        print("Nada a inserir — as três abas já estão em dia com as fontes que você tem.")
        if not a.conferir and not a.so_excel:
            print("\nMesmo assim vou conferir se a dashboard está em dia com o Excel.")
            publicar(excel, a.sem_push)
        return
    for aba, l in novos.items():
        print(f"  {aba:6s} +{len(l)} linhas")
    if a.conferir:
        print("\n[--conferir] Nada foi escrito.")
        return

    if esta_aberto(excel):
        raise SystemExit(f"\n⚠ O arquivo '{excel.name}' está ABERTO no Excel.\n"
                         "  Feche-o e rode de novo (não quero mexer num arquivo aberto).")

    # ── backup ───────────────────────────────────────────────────────────────
    BACKUP_DIR.mkdir(exist_ok=True)
    bkp = BACKUP_DIR / f"{excel.stem} — backup {datetime.now():%Y-%m-%d %H%M}{excel.suffix}"
    shutil.copy2(excel, bkp)
    print(f"\nBackup salvo: {bkp}")

    # ── escrever ─────────────────────────────────────────────────────────────
    titulo("ESCREVENDO NO EXCEL (pelo próprio Excel — gráficos e vínculos intactos)")
    pl = Planilha(excel)
    try:
        if a.refazer:
            for aba in novos:
                idx = pl.linhas_do_mes(aba, a.refazer)
                if idx:
                    print(f"  {aba}: apagando {len(idx)} linhas de {a.refazer}…")
                    pl.apagar_mes(aba, idx)
        for aba in ABAS:
            if aba in novos:
                n = pl.anexar(aba, novos[aba])
                print(f"  {aba}: {n} linhas coladas + fórmulas L:P estendidas.")
        print("  recalculando a pasta inteira e conferindo a integridade…")
        salvou = pl.fechar(salvar=True)
    except Exception:
        pl.fechar(salvar=False)
        print(f"\n⚠ Deu erro — NADA foi salvo. Seu arquivo está intacto (e há backup em {bkp}).")
        raise
    if not salvou:
        raise SystemExit(f"\nNada foi gravado. O backup intacto está em:\n  {bkp}")
    print("  Excel salvo.")

    # ── conferência ──────────────────────────────────────────────────────────
    dep = ler_estado(excel)
    print("\nDepois:")
    for aba in ABAS:
        m = sorted(dep[aba]["meses"])
        antes = sorted(est[aba]["meses"])
        seta = "  ←── novo" if m[-1] != antes[-1] else ""
        print(f"  {aba:6s} {len(m):3d} meses   {m[0]} → {m[-1]}{seta}")

    if a.so_excel:
        print("\n[--so-excel] Excel pronto. Rode com --so-dash quando quiser publicar.")
        return
    publicar(excel, a.sem_push)


if __name__ == "__main__":
    main()
